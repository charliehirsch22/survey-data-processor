"""
Data extraction utilities for the data map worksheet.

This module provides functions to extract and lookup information from the data map tab.
"""

import logging
import re

import openpyxl
from openpyxl.styles import Font

from ..constants import SHEET_DATA_MAP


def find_question_text_from_data_map(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the question text from column C of the data map in the same row as the first instance
    of the question number in column G.

    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)

    Returns:
        str: The question text from column C, or None if not found
    """
    try:
        # Search through column G starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            cell_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7

            if cell_value is not None:
                cell_str = str(cell_value).strip()

                # Check if this matches our question number
                if cell_str == str(question_number):
                    # Found the question, get column C value from same row
                    question_text = data_map_ws.cell(row=row, column=3).value  # Column C = 3
                    if question_text:
                        return str(question_text).strip()
                    else:
                        return None

        return None

    except Exception as e:
        logging.error(f"Error finding question text for question {question_number}: {e}")
        return None


def find_column_l_text_from_data_map(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the text from column L of the data map in the same row as the first instance
    of the question number in column G.

    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)

    Returns:
        str: The text from column L, or None if not found
    """
    try:
        # Search through column G starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            cell_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7

            if cell_value is not None:
                cell_str = str(cell_value).strip()

                # Check if this matches our question number
                if cell_str == str(question_number):
                    # Found the question, get column L value from same row
                    column_l_text = data_map_ws.cell(row=row, column=12).value  # Column L = 12
                    if column_l_text:
                        return str(column_l_text).strip()
                    else:
                        return None

        return None

    except Exception as e:
        logging.error(f"Error finding column L text for question {question_number}: {e}")
        return None


def find_section_number_from_data_map(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the section number from column N of the data map in the same row as the first instance
    of the question number in column G.

    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)

    Returns:
        str: The section number from column N, or None if not found
    """
    try:
        # Search through column G starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            cell_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7

            if cell_value is not None:
                cell_str = str(cell_value).strip()

                # Check if this matches our question number
                if cell_str == str(question_number):
                    # Found the question, get column N value from same row
                    section_number = data_map_ws.cell(row=row, column=14).value  # Column N = 14
                    if section_number:
                        return str(section_number).strip()
                    else:
                        return None

        return None

    except Exception as e:
        logging.error(f"Error finding section number for question {question_number}: {e}")
        return None


def find_question_column_h_text(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the first instance of a question number in column G and returns the corresponding column H value.
    Since column H contains formulas, we'll simulate the formula evaluation to get actual values.

    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)

    Returns:
        str: The evaluated value from column H, or None if not found
    """
    try:
        # Search through column G starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            cell_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7

            if cell_value is not None:
                cell_str = str(cell_value).strip()
                # Skip "System" entries
                if cell_str == "System":
                    continue

                # Check if this matches our question number
                if cell_str == str(question_number):
                    # Found the question, get column H value from same row
                    column_h_value = data_map_ws.cell(row=row, column=8).value  # Column H = 8

                    # Check if it's a formula that needs evaluation
                    if column_h_value and isinstance(column_h_value, str) and column_h_value.startswith('='):
                        # Instead of trying to evaluate complex formulas, let's get the question info from column C
                        # which contains the actual question text and information
                        question_info = data_map_ws.cell(row=row, column=3).value  # Column C
                        if question_info and str(question_info).strip():
                            return str(question_info).strip()
                        else:
                            return f"Question {question_number} data from row {row}"
                    else:
                        # Not a formula, return the value directly
                        if column_h_value is not None:
                            return str(column_h_value).strip()
                        else:
                            return None

        # Question number not found
        return None

    except Exception as e:
        logging.error(f"Error finding column H text for question {question_number}: {e}")
        return None


def find_other_specify_child_text(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the "Other Specify Child" question text from column C of the data map.

    Searches for the first row where:
    - Column G matches the question number
    - Column H contains "0, Open Text, 0, 0, 0, 0, 0, 0, 0, Other Specify Child"

    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)

    Returns:
        str: The question text from column C, or None if not found
    """
    try:
        target_h_pattern = "0, Open Text, 0, 0, , 0, 0, 0, 0, Other Specify Child"

        # Search through all rows starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            # Check column G for question number match
            column_g_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7
            # Check column H for the specific pattern
            column_h_value = data_map_ws.cell(row=row, column=8).value  # Column H = 8

            if (column_g_value is not None and
                str(column_g_value).strip() == str(question_number) and
                column_h_value is not None and
                str(column_h_value).strip() == target_h_pattern):

                # Found the matching row, get column C value (question text)
                question_text = data_map_ws.cell(row=row, column=3).value  # Column C = 3
                if question_text:
                    logging.info(f"Found Other Specify Child text at row {row} for question {question_number}")
                    return str(question_text).strip()
                else:
                    logging.warning(f"Found matching row {row} but column C is empty")
                    return None

        logging.info(f"No Other Specify Child pattern found for question {question_number}")
        return None

    except Exception as e:
        logging.error(f"Error finding Other Specify Child text for question {question_number}: {e}")
        return None


def extract_bracketed_text(text: str) -> str:
    """
    Extracts the first bracketed text from a string.

    For example: "[S1r6oe]: In which region..." returns "S1r6oe"

    Args:
        text: The text containing bracketed content

    Returns:
        str: The text inside the first brackets, or None if no brackets found
    """
    try:
        # Find the first occurrence of text within square brackets
        match = re.search(r'\[([^\]]+)\]', text)
        if match:
            return match.group(1)  # Return the content inside brackets
        else:
            return None

    except Exception as e:
        logging.error(f"Error extracting bracketed text from '{text}': {e}")
        return None


def extract_response_options(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> None:
    """
    Extracts response options from the data map and places them in the question worksheet.

    Args:
        data_map_ws: The data map worksheet
        question_ws: The question worksheet to populate
        question_number: The question number to process
    """
    try:
        logging.info(f"Extracting response options for question {question_number}")

        # Step 1: Find the section number from column N
        section_number = find_section_number_from_data_map(data_map_ws, question_number)
        if not section_number:
            logging.warning(f"No section number found for question {question_number}")
            return

        logging.info(f"Found section number for question {question_number}: {section_number}")

        # Step 2: Find all rows with "Select Option [section]" in column P
        target_pattern = f"Select Option {section_number}"
        response_rows = []

        for row in range(4, data_map_ws.max_row + 1):
            column_p_value = data_map_ws.cell(row=row, column=16).value  # Column P = 16
            if column_p_value and str(column_p_value).strip() == target_pattern:
                response_rows.append(row)

        logging.info(f"Found {len(response_rows)} response option rows for '{target_pattern}'")

        if not response_rows:
            logging.warning(f"No response options found for pattern '{target_pattern}'")
            return

        # Step 3: Extract response numbers (column D) and text (column E)
        current_row = 6  # Start placing at row 6

        for data_row in response_rows:
            # Get response number from column D
            response_number = data_map_ws.cell(row=data_row, column=4).value  # Column D = 4
            # Get response text from column E
            response_text = data_map_ws.cell(row=data_row, column=5).value  # Column E = 5

            # Place in question worksheet
            if response_number is not None:
                question_ws.cell(row=current_row, column=7, value=response_number)  # Column G = 7
            if response_text is not None:
                question_ws.cell(row=current_row, column=3, value=response_text)  # Column C = 3

            logging.info(f"Added response option {current_row-5}: {response_number} - {response_text}")
            current_row += 1

        # Step 4: Add "<>" terminator in the next row
        question_ws.cell(row=current_row, column=7, value="<>")  # Column G
        question_ws.cell(row=current_row, column=3, value="<>")  # Column C

        logging.info(f"Added terminator '<>' at row {current_row}")

        # Step 5: Add COUNTIFS formula to column D starting at row 6
        formula = '=COUNTIFS(OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($G$4, \'raw data\'!$C$2:$AJC$2, 0)-1), $G6, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($I6, \'raw data\'!$C$2:$AJC$2, 0)-1), $J6, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($K6, \'raw data\'!$C$2:$AJC$2, 0)-1), $L6, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($M6, \'raw data\'!$C$2:$AJC$2, 0)-1), $N6)'

        # Add formula to D6 and drag down to include the terminator row (current_row has "<>")
        for formula_row in range(6, current_row + 1):  # Include the "<>" row
            # Adjust the formula for the current row by replacing "6" with the current row number
            adjusted_formula = formula.replace('$G6', f'$G{formula_row}').replace('$I6', f'$I{formula_row}').replace('$J6', f'$J{formula_row}').replace('$K6', f'$K{formula_row}').replace('$L6', f'$L{formula_row}').replace('$M6', f'$M{formula_row}').replace('$N6', f'$N{formula_row}')
            question_ws.cell(row=formula_row, column=4, value=adjusted_formula)  # Column D = 4

        # Step 6: Add "record" and "<>" pattern to columns I:N starting at row 6 with blue font
        pattern_values = ["record", "<>", "record", "<>", "record", "<>"]  # I, J, K, L, M, N
        blue_font = Font(color="0000FF")

        for row_num in range(6, current_row + 1):  # Include the "<>" row
            for col_index, value in enumerate(pattern_values):
                col_num = 9 + col_index  # Column I = 9, J = 10, K = 11, L = 12, M = 13, N = 14
                question_ws.cell(row=row_num, column=col_num, value=value).font = blue_font

        # Step 7: Add percentage formula to column E starting at row 6
        # The denominator is the absolute reference to the cell in column D next to the "<>" in column C
        denominator_row = current_row  # This is the row with "<>" in column C

        for row_num in range(6, current_row + 1):  # Include the "<>" row
            # Create percentage formula: relative numerator / absolute denominator
            percentage_formula = f"=D{row_num}/$D${denominator_row}"
            cell = question_ws.cell(row=row_num, column=5, value=percentage_formula)  # Column E = 5
            # Format as percentage
            cell.number_format = '0.0%'

        logging.info(f"Added COUNTIFS formula to column D from row 6 to {current_row}")
        logging.info(f"Added record/<> pattern to columns I:N from row 6 to {current_row} with blue font")
        logging.info(f"Added percentage formula to column E from row 6 to {current_row} with denominator $D${denominator_row}")
        logging.info(f"Successfully extracted {len(response_rows)} response options for question {question_number}")

    except Exception as e:
        logging.error(f"Error extracting response options for question {question_number}: {e}")
        raise
