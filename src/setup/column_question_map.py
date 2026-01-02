"""
Column question map worksheet setup.

This module handles initial setup and configuration of the column question map tab.
"""

import logging
import re

import openpyxl
from openpyxl.styles import Alignment

from ..constants import (
    SHEET_COLUMN_QUESTION_MAP,
    SHEET_RAW_DATA,
    COL_WIDTH_NARROW,
    COL_WIDTH_WIDE,
)
from ..formatters.styles import create_thin_border, create_pale_blue_fill


def column_question_map_initial_setup(workbook: openpyxl.Workbook) -> None:
    """
    Performs initial setup on the column question map tab.

    Args:
        workbook (openpyxl.Workbook): The workbook containing the column question map tab.
    """
    try:
        logging.info("Starting column question map initial setup...")

        # Get the column question map worksheet
        if SHEET_COLUMN_QUESTION_MAP not in [ws.title for ws in workbook.worksheets]:
            logging.warning(f"No '{SHEET_COLUMN_QUESTION_MAP}' tab found, skipping column question map setup")
            return

        ws = workbook[SHEET_COLUMN_QUESTION_MAP]

        # Remove gridlines from the worksheet
        ws.sheet_view.showGridLines = False
        logging.info("Removed gridlines from column question map worksheet")

        # Set column widths
        ws.column_dimensions['A'].width = COL_WIDTH_NARROW
        ws.column_dimensions['B'].width = COL_WIDTH_NARROW
        ws.column_dimensions['C'].width = COL_WIDTH_WIDE
        ws.column_dimensions['D'].width = COL_WIDTH_WIDE
        ws.column_dimensions['E'].width = COL_WIDTH_WIDE
        ws.column_dimensions['F'].width = COL_WIDTH_WIDE
        ws.column_dimensions['G'].width = COL_WIDTH_WIDE
        ws.column_dimensions['H'].width = COL_WIDTH_WIDE
        logging.info(f"Set column widths: A:B={COL_WIDTH_NARROW}, C:H={COL_WIDTH_WIDE}")

        # Add column headers in row 2, columns C through H
        headers = [
            "All question columns",
            "System or Survey",
            "Question markers",
            "Question Number",
            "Unique question markers",
            "Question Number Map"
        ]

        for i, header in enumerate(headers):
            col = i + 3  # C=3, D=4, E=5, F=6, G=7, H=8
            ws.cell(row=2, column=col, value=header)

        logging.info("Added column headers in row 2: C2:H2")

        # Format headers in row 2 with borders and pale blue background
        pale_blue_fill = create_pale_blue_fill()
        thin_border = create_thin_border()
        center_alignment = Alignment(horizontal='center', vertical='center')

        headers_formatted = 0
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            if cell.value is not None:  # Only format cells with content
                cell.border = thin_border
                cell.fill = pale_blue_fill
                cell.alignment = center_alignment
                headers_formatted += 1

        logging.info(f"Formatted {headers_formatted} column headers in row 2 with borders, pale blue background, and center alignment")

        # Copy column headers from raw data tab and transpose to column B
        if SHEET_RAW_DATA in [ws_temp.title for ws_temp in workbook.worksheets]:
            raw_data_ws = workbook[SHEET_RAW_DATA]

            # Find the last column with text in row 2 of raw data
            last_col_with_text = 1
            for col in range(1, raw_data_ws.max_column + 1):
                if raw_data_ws.cell(row=2, column=col).value is not None and str(raw_data_ws.cell(row=2, column=col).value).strip():
                    last_col_with_text = col

            # Copy headers from C2 onwards in raw data tab
            headers_copied = 0
            current_row = 3  # Start at C3 in column question map

            for col in range(3, last_col_with_text + 1):  # Start from C2 (column 3)
                header_value = raw_data_ws.cell(row=2, column=col).value
                if header_value is not None:
                    # Paste as value (transposed) into column C of column question map
                    ws.cell(row=current_row, column=3, value=header_value)  # Column C = 3
                    headers_copied += 1
                    current_row += 1

            logging.info(f"Copied and transposed {headers_copied} column headers from raw data tab to column C (C3 onwards)")
        else:
            logging.warning("Raw data tab not found, skipping header copying")

        # Add formulas starting in row 3
        # D3 formula
        ws['D3'] = '=IF(ISTEXT(LEFT(C3,1)),IF(EXACT(LEFT(C3,1),UPPER(LEFT(C3,1))),"Survey","System"),"First Char not Letter")'
        # E3 formula
        ws['E3'] = '=IF(D3="System","System",IF(ISNUMBER(FIND("_",C3)),LEFT(C3,FIND("_",C3)-1),IF(ISNUMBER(FIND("none",C3)),LEFT(C3,FIND("none",C3)-1),IF(ISNUMBER(FIND("r",C3)),LEFT(C3,FIND("r",C3)-1),C3))))'
        # F3 formula
        ws['F3'] = '=IFERROR(INDEX($H$3:$H$200, MATCH($E3, $G$3:$G$200, 0)), "System")'
        logging.info("Added formulas to D3, E3, and F3")

        # Fill column H with sequential numbers 1-200 starting from H3
        for i in range(1, 201):  # 1 to 200
            row_num = i + 2  # H3 starts at row 3, so H3=1, H4=2, etc.
            ws.cell(row=row_num, column=8, value=i)  # Column H = 8

        logging.info("Added sequential numbers 1-200 in column H starting at H3")

        # Copy formulas D3:G3 down to the last row with text in column C
        # Find the last row with text in column C
        last_row_with_text = 1
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=3).value is not None and str(ws.cell(row=row, column=3).value).strip():
                last_row_with_text = row

        if last_row_with_text > 3:  # Only copy if there are rows below row 3
            # Copy formulas from D3:G3 down to the last row with text
            for row in range(4, last_row_with_text + 1):  # Start from row 4 (next row after formulas)
                for col in range(4, 8):  # D=4, E=5, F=6, G=7
                    source_cell = ws.cell(row=3, column=col)
                    target_cell = ws.cell(row=row, column=col)

                    # Copy the formula, adjusting relative references
                    if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                        # Manually adjust relative row references in the formula
                        formula = source_cell.value
                        row_offset = row - 3

                        # Replace relative row references (e.g., C3 -> C4, C5, etc.)
                        def replace_row_ref(match):
                            col_ref = match.group(1)
                            row_ref = int(match.group(2))
                            # Only adjust relative references (not absolute ones with $)
                            if '$' not in match.group(0):
                                new_row = row_ref + row_offset
                                return f"{col_ref}{new_row}"
                            return match.group(0)

                        # Pattern to match column+row references (e.g., C3, D3, etc.)
                        adjusted_formula = re.sub(r'([A-Z]+)(\d+)', replace_row_ref, formula)
                        target_cell.value = adjusted_formula

            copied_rows = last_row_with_text - 3
            logging.info(f"Copied formulas from D3:G3 down to row {last_row_with_text} ({copied_rows} additional rows)")
        else:
            logging.info("No additional rows to copy formulas to")

        logging.info(f"Found last row with text in column C: row {last_row_with_text}")

        # After formulas are copied, simulate the E column results to populate G with unique values
        # Since we can't evaluate Excel formulas in openpyxl, we'll simulate the E3 formula logic
        unique_values = []  # Use list to preserve order
        seen_values = set()  # Track what we've seen to avoid duplicates

        # Process each column header to simulate what column E formulas would produce
        for row in range(3, last_row_with_text + 1):
            # Get the column header from column C
            header_value = ws.cell(row=row, column=3).value  # Column C = 3
            if header_value is not None:
                header_str = str(header_value).strip()

                # Simulate the E3 formula logic: question marker extraction
                # Check if first character is a letter and uppercase (Survey vs System)
                if header_str and len(header_str) > 0:
                    first_char = header_str[0]
                    if first_char.isalpha() and first_char.isupper():
                        # This would be "Survey" - extract question marker
                        question_marker = header_str

                        # Apply the same logic as E3 formula
                        if "_" in question_marker:
                            question_marker = question_marker[:question_marker.find("_")]
                        elif "none" in question_marker:
                            question_marker = question_marker[:question_marker.find("none")]
                        elif "r" in question_marker:
                            question_marker = question_marker[:question_marker.find("r")]

                        # Add to unique list if not "System", not empty, and not already seen
                        if question_marker and question_marker != "System" and question_marker not in seen_values:
                            unique_values.append(question_marker)
                            seen_values.add(question_marker)

        # unique_values list now maintains the order from column E
        unique_list = unique_values

        # Populate column G starting at G3 with unique values
        for i, unique_value in enumerate(unique_list):
            row_num = i + 3  # Start at G3
            ws.cell(row=row_num, column=7, value=unique_value)  # Column G = 7

        logging.info(f"Populated column G with {len(unique_list)} unique question markers simulated from column C headers")

        # Apply center alignment to columns F, G, and H
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Apply center alignment to all of columns F, G, H (broader range to ensure coverage)
        for col in range(6, 9):  # Columns F=6, G=7, H=8
            for row in range(1, 1000):  # Apply to a large range to ensure all cells are covered
                ws.cell(row=row, column=col).alignment = center_alignment

        logging.info("Applied center alignment to entire columns F, G, and H")

        # Column G setup will be handled separately per user instructions

        logging.info("Column question map initial setup completed successfully")

    except Exception as e:
        logging.error(f"Error during column question map initial setup: {e}")
        raise
