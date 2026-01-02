"""
Raw data worksheet setup.

This module handles initial setup and formatting of the raw data tab.
"""

import logging

import openpyxl
from openpyxl.styles import Alignment

from ..constants import SHEET_RAW_DATA, COL_WIDTH_NARROW
from ..formatters.styles import create_thin_border, create_pale_blue_fill


def raw_data_initial_setup(workbook: openpyxl.Workbook) -> None:
    """
    Performs initial setup on the raw data tab:
    - Inserts 2 columns at the front with width 3
    - Inserts 1 row at the top
    - Formats column headers (row 2) with borders and light blue background

    Args:
        workbook (openpyxl.Workbook): The workbook containing the raw data tab.
    """
    try:
        logging.info("Starting raw data initial setup...")

        # Get the raw data worksheet
        if SHEET_RAW_DATA not in [ws.title for ws in workbook.worksheets]:
            logging.warning(f"No '{SHEET_RAW_DATA}' tab found, skipping raw data setup")
            return

        ws = workbook[SHEET_RAW_DATA]

        # Insert 2 columns at the front
        ws.insert_cols(1, 2)
        logging.info("Inserted 2 columns at the front")

        # Set width of the first 2 columns to 3
        ws.column_dimensions['A'].width = COL_WIDTH_NARROW
        ws.column_dimensions['B'].width = COL_WIDTH_NARROW
        logging.info(f"Set width of first 2 columns to {COL_WIDTH_NARROW}")

        # Insert 1 row at the top
        ws.insert_rows(1, 1)
        logging.info("Inserted 1 row at the top")

        # Remove gridlines from the worksheet
        ws.sheet_view.showGridLines = False
        logging.info("Removed gridlines from worksheet")

        # Format column headers in row 2
        # Find the actual header row (should be row 2 after inserting 1 row)
        # But let's check both row 1 and 2 to be safe
        header_row = 2

        # Check if row 2 has headers, if not try row 1
        has_headers_row2 = any(ws.cell(row=2, column=col).value is not None for col in range(3, ws.max_column + 1))
        if not has_headers_row2:
            has_headers_row1 = any(ws.cell(row=1, column=col).value is not None for col in range(3, ws.max_column + 1))
            if has_headers_row1:
                header_row = 1

        logging.info(f"Using row {header_row} as header row")

        # Find the last column with data in the header row
        last_col = ws.max_column
        for col in range(3, ws.max_column + 1):  # Start from column 3 (C) since we inserted 2 columns
            if ws.cell(row=header_row, column=col).value is None:
                last_col = col - 1
                break

        # Create formatting styles
        thin_border = create_thin_border()
        pale_blue_fill = create_pale_blue_fill()
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Apply formatting to all header cells in the header row
        headers_formatted = 0
        for col in range(3, last_col + 1):  # Start from column 3 (C) since we inserted 2 columns
            cell = ws.cell(row=header_row, column=col)
            if cell.value is not None:  # Only format cells with content
                cell.border = thin_border
                cell.fill = pale_blue_fill
                cell.alignment = center_alignment
                headers_formatted += 1

        logging.info(f"Formatted {headers_formatted} column headers with borders, pale blue background, and center alignment")
        logging.info("Raw data initial setup completed successfully")

    except Exception as e:
        logging.error(f"Error during raw data initial setup: {e}")
        raise
