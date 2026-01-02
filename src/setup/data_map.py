"""
Data map worksheet setup.

This module handles initial setup and formula configuration of the data map tab.
"""

import logging
import re

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from ..constants import (
    SHEET_DATA_MAP,
    COL_WIDTH_NARROW,
    COL_WIDTH_STANDARD,
    COL_WIDTH_EXTRA_WIDE,
)
from ..formatters.styles import create_thin_border, create_pale_blue_fill


def data_map_initial_setup(workbook: openpyxl.Workbook) -> None:
    """
    Performs initial setup on the data map tab:
    - Removes gridlines
    - Adds 2 columns on the left with width 3
    - Adds 3 rows at the top
    - Adds headers in row 2: "Question Info", "System Response Option", "Text Response Option", "Question Number"
    - Makes column F width 3

    Args:
        workbook (openpyxl.Workbook): The workbook containing the data map tab.
    """
    try:
        logging.info("Starting data map initial setup...")

        # Get the data map worksheet
        if SHEET_DATA_MAP not in [ws.title for ws in workbook.worksheets]:
            logging.warning(f"No '{SHEET_DATA_MAP}' tab found, skipping data map setup")
            return

        ws = workbook[SHEET_DATA_MAP]

        # Remove gridlines from the worksheet
        ws.sheet_view.showGridLines = False
        logging.info("Removed gridlines from data map worksheet")

        # Insert 2 columns at the front
        ws.insert_cols(1, 2)
        logging.info("Inserted 2 columns at the front")

        # Set width of the first 2 columns to 3
        ws.column_dimensions['A'].width = COL_WIDTH_NARROW
        ws.column_dimensions['B'].width = COL_WIDTH_NARROW
        logging.info(f"Set width of first 2 columns to {COL_WIDTH_NARROW}")

        # Set width of column F to 3
        ws.column_dimensions['F'].width = COL_WIDTH_NARROW
        logging.info(f"Set width of column F to {COL_WIDTH_NARROW}")

        # Set column widths for data map layout
        ws.column_dimensions['C'].width = COL_WIDTH_EXTRA_WIDE
        logging.info(f"Set width of column C to {COL_WIDTH_EXTRA_WIDE}")

        # Set width of columns D and E to 13
        ws.column_dimensions['D'].width = COL_WIDTH_STANDARD
        ws.column_dimensions['E'].width = COL_WIDTH_STANDARD
        logging.info(f"Set width of columns D:E to {COL_WIDTH_STANDARD}")

        # Set width of columns G through Z to 13
        for col_letter in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
            ws.column_dimensions[col_letter].width = COL_WIDTH_STANDARD
        logging.info(f"Set width of columns G:Z to {COL_WIDTH_STANDARD}")

        # Insert 3 rows at the top
        ws.insert_rows(1, 3)
        logging.info("Inserted 3 rows at the top")

        # Add headers in row 2
        ws['C2'] = "Question Info"
        ws['D2'] = "System Response Option"
        ws['E2'] = "Text Response Option"
        ws['G2'] = "Question Number"

        # Add additional headers from H to AG
        headers_h_to_ag = [
            "Question Type", "Question Text", "Question Code", "System or Survey Q",
            "Question Prefix", "Section Marker", "Section Number", "Row Tag",
            "Row Tag + Section", "Other Text Entry Flag", "Other Question Type Flag",
            "Other Question Type", "System Question Flag", "System Question",
            "Open Text", "Numerical", "Simple Select", "Multi Select (placeholder)",
            "Rank", "Matrix", "Loop", "Double Loop", "Double Loop Flag #0",
            "Double Loop Flag #1", "Double Loop Flag #2", "Double Loop Flag #3"
        ]

        # Start from column H (8) and add each header
        for i, header in enumerate(headers_h_to_ag):
            col_num = 8 + i  # H is column 8
            if col_num <= 33:  # AG is column 33
                ws.cell(row=2, column=col_num, value=header)

        logging.info("Added headers in row 2: Question Info, System Response Option, Text Response Option, Question Number, and additional headers H:AG")

        # Set text wrapping for column C and row 2
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        for cell in ws['C']:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        logging.info("Applied text wrapping to column C and row 2")

        # Format headers in row 2 with borders and pale blue background
        pale_blue_fill = create_pale_blue_fill()
        thin_border = create_thin_border()

        headers_formatted = 0
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            if cell.value is not None:  # Only format cells with content
                cell.border = thin_border
                cell.fill = pale_blue_fill
                # Preserve the wrap_text alignment that was set earlier
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                headers_formatted += 1

        logging.info(f"Formatted {headers_formatted} column headers in row 2 with borders, pale blue background, and center alignment")

        # Auto-fit row height for row 2 to accommodate wrapped text
        # Calculate the maximum lines needed based on text content and column width
        max_lines = 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            if cell.value is not None:
                text = str(cell.value)
                # Get column width (default to 13 if not set)
                col_letter = get_column_letter(col)
                col_width = ws.column_dimensions[col_letter].width or 13

                # Estimate characters per line based on column width
                chars_per_line = max(1, int(col_width * 0.8))  # Rough estimate
                lines_needed = max(1, len(text) // chars_per_line + (1 if len(text) % chars_per_line > 0 else 0))

                # Account for word wrapping (longer words need more lines)
                words = text.split()
                if words:
                    longest_word = max(len(word) for word in words)
                    if longest_word > chars_per_line:
                        lines_needed = max(lines_needed, 2)

                max_lines = max(max_lines, lines_needed)

        # Set row height based on estimated lines (Excel default line height is about 15 points)
        estimated_height = max_lines * 15 + 5  # Add some padding
        ws.row_dimensions[2].height = estimated_height
        logging.info(f"Set row 2 height to {estimated_height} points to accommodate {max_lines} lines of wrapped text")

        # Add formulas starting in row 4
        # G4 formula
        ws['G4'] = "=INDEX('column question map'!$F$3:$F$939, MATCH('data map'!$L4, 'column question map'!$E$3:$E$939, 0))"
        # H4 formula
        ws['H4'] = '=U4&", "&V4&", "&W4&", "&X4&", "&Y4&", "&Z4&", "&AA4&", "&AB4&", "&AC4&", "&S4'
        # I4 formula
        ws['I4'] = '=INDEX($C$4:$C$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0))'
        # J4 formula
        ws['J4'] = '=SUBSTITUTE(SUBSTITUTE(LEFT(I4,FIND(":",I4)-1),"[",""),"]","")'
        # K4 formula
        ws['K4'] = '=IF(ISTEXT(LEFT(J4,1)),IF(EXACT(LEFT(J4,1),UPPER(LEFT(J4,1))),"Survey","System"),"First Char not Letter")'
        # L4 formula
        ws['L4'] = '=IF(K4="System","System",IF(ISNUMBER(FIND("_",J4)),LEFT(J4,FIND("_",J4)-1),IF(ISNUMBER(FIND("none",J4)),LEFT(J4,FIND("none",J4)-1),IF(ISNUMBER(FIND("r",J4)),LEFT(J4,FIND("r",J4)-1),J4))))'
        # M4 formula
        ws['M4'] = '=IF(AND(C4="", D4="", E4=""), "End", IF(AND(OFFSET(C4,-1,0)="", OFFSET(D4,-1,0)="", OFFSET(E4,-1,0)=""), "Start", "Mid"))'
        # N4 formula
        ws['N4'] = '=COUNTIFS($M$4:M4, "Start")'
        # O4 formula
        ws['O4'] = '=IF(M4="Start","Question Text",IF(O3="Question Text","Response Type",IF(ISNUMBER(D4),"Select Option",IF(AND(LEFT(D4,1)="[",RIGHT(D4,1)="]"),"Bracketed Sub-Question","End"))))'
        # P4 formula
        ws['P4'] = '=O4 & " " &N4'
        # Q4 formula
        ws['Q4'] = '=IF(ISNUMBER(SEARCH("oe]", C4)), "Other Text Entry", 0)'
        # R4 formula
        ws['R4'] = '=IFERROR(IF(Q4="Other Text Entry","Other Specify Child",IF(INDEX($Q$4:$Q$3200,MATCH(O4&" "&TEXT(N4+1,"0"),$P$4:$P$3200,0))="Other Text Entry","Other Specify Parent",0)),0)'
        # S4 formula
        ws['S4'] = '=INDEX($R$4:$R$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0))'
        # T4 formula
        ws['T4'] = '=IF(OR(ISNUMBER(SEARCH("[record]",C4)),ISNUMBER(SEARCH("[uuid]",C4)),ISNUMBER(SEARCH("[date]",C4)),ISNUMBER(SEARCH("[markers]",C4)),ISNUMBER(SEARCH("[status]",C4)),ISNUMBER(SEARCH("conditions: Conditions",C4)),ISNUMBER(SEARCH("[vlist]",C4)),ISNUMBER(SEARCH("[qtime]",C4)),ISNUMBER(SEARCH("[vos]",C4)),ISNUMBER(SEARCH("[vosr15oe]",C4)),ISNUMBER(SEARCH("[vbrowser]",C4)),ISNUMBER(SEARCH("[vbrowser15oe]",C4)),ISNUMBER(SEARCH("[vmobiledevice]",C4)),ISNUMBER(SEARCH("[vmobileos]",C4)),ISNUMBER(SEARCH("[start_date]",C4)),ISNUMBER(SEARCH("[vdropout]",C4)),ISNUMBER(SEARCH("[source]",C4)),ISNUMBER(SEARCH("[decLang]",C4)),ISNUMBER(SEARCH("[list]",C4)),ISNUMBER(SEARCH("[userAgent]",C4)),ISNUMBER(SEARCH("[fp_etag]",C4)),ISNUMBER(SEARCH("[fp_html5]",C4)),ISNUMBER(SEARCH("[fp_flash]",C4)),ISNUMBER(SEARCH("[fp_browser]",C4)),ISNUMBER(SEARCH("[dcua]",C4)),ISNUMBER(SEARCH("[url]",C4)),ISNUMBER(SEARCH("[session]",C4)),ISNUMBER(SEARCH("[s24627]",C4)),ISNUMBER(SEARCH("[s25023]",C4))),"System Question",0)'
        # U4 formula
        ws['U4'] = '=INDEX($T$4:$T$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0))'
        # V4 formula
        ws['V4'] = '=IF(INDEX($C$4:$C$3200, MATCH("Response Type "&N4, $P$4:$P$3200, 0)) = "Open text response", "Open Text", 0)'
        # W4 formula
        ws['W4'] = '=IF(INDEX($C$4:$C$3200, MATCH("Response Type "&N4, $P$4:$P$3200, 0)) = "Open numeric response", "Numerical", IF(AND(LEFT(INDEX($C$4:$C$3200, MATCH("Response Type "&N4, $P$4:$P$3200, 0)), 5) = "Value", ISERROR(MATCH("Select Option "&N4, $P$4:$P$3200, 0))), "Numerical", 0))'
        # X4 formula
        ws['X4'] = '=IF(AND(NOT(S4="Other Specify Child"), T4=0, U4=0, V4=0, W4=0, LEFT(INDEX($C$4:$C$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0)), 1) = "["), "Simple Select", 0)'
        # Y4 formula
        ws['Y4'] = '0'
        # Z4 formula
        ws['Z4'] = '=IF(ISNUMBER(SEARCH("rank", INDEX($C$4:$C$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0)))), "Rank", 0)'
        # AA4 formula
        ws['AA4'] = '=IF(AND(Z4=0, NOT((ISERROR(MATCH("Bracketed Sub-Question " & N4, $P$4:$P$3200, 0))))), "Matrix", 0)'
        # AB4 formula
        ws['AB4'] = '=IF(ISNUMBER(SEARCH("_", INDEX($C$4:$C$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0)))), "Loop", 0)'
        # AC4 formula
        ws['AC4'] = '=INDEX($AD$4:$AD$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0))'
        # AD4 formula
        ws['AD4'] = '=IF(AE4*AF4*AG4=1, "Double Loop", 0)'
        # AE4 formula
        ws['AE4'] = '=IFERROR(IF(AND(ISNUMBER(SEARCH("_",C4)),ISERROR(VALUE(MID(C4,SEARCH("_",C4)-1,1))),MID(C4,SEARCH("_",C4)-1,1)<>""),1,0), 0)'
        # AF4 formula
        ws['AF4'] = '=IF(AND(ISNUMBER(SEARCH("Lr",C4)),ISNUMBER(VALUE(MID(C4,SEARCH("Lr",C4)+2,1)))),1,0)'
        # AG4 formula
        ws['AG4'] = '=IF(OR(ISNUMBER(SEARCH("0r",C4)),ISNUMBER(SEARCH("1r",C4)),ISNUMBER(SEARCH("2r",C4)),ISNUMBER(SEARCH("3r",C4)),ISNUMBER(SEARCH("4r",C4)),ISNUMBER(SEARCH("5r",C4)),ISNUMBER(SEARCH("6r",C4)),ISNUMBER(SEARCH("7r",C4)),ISNUMBER(SEARCH("8r",C4)),ISNUMBER(SEARCH("9r",C4))),1,0)'
        logging.info("Added formulas to G4 through AG4 (27 total formulas)")

        # Find the last row with text in column C
        last_row_with_text = 1
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=3).value is not None and str(ws.cell(row=row, column=3).value).strip():
                last_row_with_text = row

        # Copy formulas down to one row below the last row with text in column C
        target_last_row = last_row_with_text + 1

        if target_last_row > 4:  # Only copy if there are rows to copy to
            # Copy formulas from G4:AG4 down to the target range (including G column)

            # Define the source range (G4:AG4) - including G column
            source_range = f"G4:AG4"
            # Define the target range (G5:AG{target_last_row}) - including G column
            target_range = f"G5:AG{target_last_row}"

            # Get source cells
            source_cells = ws[source_range]

            # Copy formulas to each row
            for row_idx, target_row in enumerate(range(5, target_last_row + 1)):
                for col_idx, source_cell in enumerate(source_cells[0]):  # source_cells[0] since it's a single row
                    target_cell = ws.cell(row=target_row, column=source_cell.column)

                    if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                        # Copy the formula - openpyxl will automatically adjust relative references
                        target_cell.value = source_cell.value

                        # Manually adjust relative row references in the formula
                        formula = source_cell.value
                        row_offset = target_row - 4

                        # Replace relative row references (e.g., C4 -> C5, C6, etc.)
                        def replace_row_ref(match):
                            col_ref = match.group(1)
                            row_ref = int(match.group(2))
                            # Only adjust relative references (not absolute ones with $)
                            if '$' not in match.group(0):
                                new_row = row_ref + row_offset
                                return f"{col_ref}{new_row}"
                            return match.group(0)

                        # Pattern to match column+row references (e.g., C4, M4, etc.)
                        adjusted_formula = re.sub(r'([A-Z]+)(\d+)', replace_row_ref, formula)
                        target_cell.value = adjusted_formula

            copied_rows = target_last_row - 4
            logging.info(f"Copied formulas from G4:AG4 down to row {target_last_row} ({copied_rows} additional rows)")
        else:
            logging.info("No additional rows to copy formulas to")

        logging.info(f"Found last row with text in column C: row {last_row_with_text}")
        logging.info("Data map initial setup completed successfully")

    except Exception as e:
        logging.error(f"Error during data map initial setup: {e}")
        raise
