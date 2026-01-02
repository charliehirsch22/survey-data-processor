"""
Initial setup for the workbook.

This module handles renaming tabs and creating blank worksheets.
"""

import logging

import openpyxl

from ..constants import (
    SHEET_RAW_DATA,
    SHEET_DATA_MAP,
    SHEET_COLUMN_QUESTION_MAP,
    SHEET_LOOP_VARIABLES,
)


def initial_set_up(workbook: openpyxl.Workbook) -> None:
    """
    Performs initial setup on the workbook:
    - Renames tabs "A", "B", or "raw data" to "raw data"
    - Renames tabs like "datamap" to "data map"
    - Adds blank tabs "column question map" and "loop variables"

    Args:
        workbook (openpyxl.Workbook): The workbook to set up.
    """
    try:
        logging.info("Starting initial setup...")

        # Get all worksheet names
        sheet_names = workbook.sheetnames
        logging.info(f"Found sheets: {sheet_names}")

        # Step 1: Rename tabs A, B, A1, or "raw data" to "raw data"
        for sheet_name in sheet_names:
            sheet_lower = sheet_name.lower()
            if (sheet_lower in ['a', 'b', SHEET_RAW_DATA] or
                sheet_lower.startswith('a') and len(sheet_lower) <= 3):
                if sheet_lower != SHEET_RAW_DATA:
                    workbook[sheet_name].title = SHEET_RAW_DATA
                    logging.info(f"Renamed '{sheet_name}' to '{SHEET_RAW_DATA}'")
                break

        # Step 2: Rename datamap-like tabs to "data map"
        for sheet_name in workbook.sheetnames:
            if 'datamap' in sheet_name.lower():
                workbook[sheet_name].title = SHEET_DATA_MAP
                logging.info(f"Renamed '{sheet_name}' to '{SHEET_DATA_MAP}'")
                break

        # Step 3: Add new blank tabs if they don't exist
        required_tabs = [SHEET_COLUMN_QUESTION_MAP, SHEET_LOOP_VARIABLES]
        existing_tabs = [ws.title.lower() for ws in workbook.worksheets]

        for tab_name in required_tabs:
            if tab_name.lower() not in existing_tabs:
                new_sheet = workbook.create_sheet(title=tab_name)
                logging.info(f"Created new tab: '{tab_name}'")
            else:
                logging.info(f"Tab '{tab_name}' already exists")

        logging.info("Initial setup completed successfully")

    except Exception as e:
        logging.error(f"Error during initial setup: {e}")
        raise
