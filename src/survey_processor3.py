#!/usr/bin/env python3
"""
Survey Data Processor

Converts raw survey data files into formatted Excel workbooks.

"""

import pandas as pd
import numpy as np
from pathlib import Path
import chardet
import logging
from typing import Dict, List, Optional, Tuple
import json
import re
from enum import Enum
import openpyxl
from openpyxl.utils import get_column_letter

def load_raw_survey(file_path: str) -> openpyxl.Workbook:
    """
    Loads an Excel file and returns an openpyxl Workbook object for editing.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        openpyxl.Workbook: The loaded workbook.
    
    Raises:
        FileNotFoundError: If the file doesn't exist.
        ValueError: If the file is not a valid Excel file.
    """
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if file_path_obj.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
            raise ValueError(f"Not a valid Excel file: {file_path}")
        
        return openpyxl.load_workbook(file_path)
    except Exception as e:
        logging.error(f"Error loading workbook {file_path}: {e}")
        raise

def rename_datamap_tab(workbook: openpyxl.Workbook) -> None:
    """
    Renames any worksheet with a name similar to 'datamap' to 'data map'.

    Args:
        workbook (openpyxl.Workbook): The workbook to edit.
    """
    for ws in workbook.worksheets:
        if re.search(r'data\s*map|datamap', ws.title, re.IGNORECASE):
            ws.title = 'data map'

def rename_rawdata_tab(workbook: openpyxl.Workbook) -> None:
    """
    Renames any worksheet with a name similar to 'A1' or 'raw data' to 'raw data'.

    Args:
        workbook (openpyxl.Workbook): The workbook to edit.
    """
    for ws in workbook.worksheets:
        if re.search(r'^A1$', ws.title, re.IGNORECASE) or re.search(r'raw\s*data', ws.title, re.IGNORECASE):
            ws.title = 'raw data'

def shift_worksheet_to_C2(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Ensures worksheet has blank first row and first two columns (A, B) with width 3,
    without removing any existing data. Data is shifted if necessary.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to modify.
    """
    # Check if first row has any data
    first_row_has_data = any(cell.value is not None for cell in ws[1])
    
    # Check if columns A or B have any data
    col_a_has_data = any(cell.value is not None for cell in ws['A'])
    col_b_has_data = any(cell.value is not None for cell in ws['B'])
    
    # Insert blank row at top if first row has data
    if first_row_has_data:
        ws.insert_rows(1)
    
    # Insert blank columns if A or B have data
    cols_to_insert = 0
    if col_b_has_data:
        cols_to_insert = 2
    elif col_a_has_data:
        cols_to_insert = 1
    
    if cols_to_insert > 0:
        ws.insert_cols(1, amount=cols_to_insert)
    
    # Always set width of columns A and B to 3
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 3
    
    # Remove gridlines
    ws.sheet_view.showGridLines = False

def process_raw_data_tab(workbook: openpyxl.Workbook) -> None:
    """
    Processes the 'raw data' tab by applying shift_worksheet_to_C2.

    Args:
        workbook (openpyxl.Workbook): The workbook containing the 'raw data' tab.
    """
    try:
        raw_data_ws = workbook['raw data']
        shift_worksheet_to_C2(raw_data_ws)
    except KeyError:
        logging.warning("No 'raw data' tab found in workbook")

def process_data_map_tab(workbook: openpyxl.Workbook) -> None:
    """
    Processes the 'data map' tab by applying minimal formatting and adding headers only.

    Args:
        workbook (openpyxl.Workbook): The workbook containing the 'data map' tab.
    """
    try:
        data_map_ws = workbook['data map']
        
        # Apply the standard formatting first (this handles data preservation)
        shift_worksheet_to_C2(data_map_ws)
        
        # Add column headers to row 2 ONLY
        data_map_ws['C2'] = "Question Info"
        data_map_ws['D2'] = "Code"
        data_map_ws['E2'] = "Text Map"
        data_map_ws['F2'] = "Helper #1"
        data_map_ws['G2'] = "Question Tag #1"
        data_map_ws['H2'] = "Question Tag #2"
        data_map_ws['I2'] = "Question Tag #3"
        data_map_ws['J2'] = "Question Tag #4"
        data_map_ws['K2'] = "Other Flag"
        
        # Format column C: width 45 and text wrap
        data_map_ws.column_dimensions['C'].width = 45
        for row in data_map_ws.iter_rows(min_col=3, max_col=3):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        # Center align headers in row 2 for columns D through K
        for col_num in range(4, 12):  # Columns D through K
            cell = data_map_ws.cell(row=2, column=col_num)
            if cell.value:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # STOP HERE - don't do any other processing that might corrupt the data
        
    except KeyError:
        logging.warning("No 'data map' tab found in workbook")

def save_workbook(workbook: openpyxl.Workbook, file_path: str) -> None:
    """
    Saves the workbook to the specified file path.

    Args:
        workbook (openpyxl.Workbook): The workbook to save.
        file_path (str): Path where to save the workbook.
    """
    try:
        workbook.save(file_path)
        logging.info(f"Workbook saved successfully to: {file_path}")
    except Exception as e:
        logging.error(f"Error saving workbook to {file_path}: {e}")
        raise

def process_survey_file(input_file: str, output_file: str = None) -> None:
    """
    Main function to process a survey file.

    Args:
        input_file (str): Path to the input Excel file.
        output_file (str, optional): Path for the output file. If None, input file is modified in place.
    """
    try:
        # Load the workbook
        workbook = load_raw_survey(input_file)
        
        # Rename tabs to standard names
        rename_datamap_tab(workbook)
        rename_rawdata_tab(workbook)
        
        # Process the tabs
        process_raw_data_tab(workbook)
        process_data_map_tab(workbook)
        
        # Determine output path
        if output_file:
            save_workbook(workbook, output_file)
            print(f"Successfully processed survey file: {input_file}")
            print(f"Output saved to: {output_file}")
        else:
            save_workbook(workbook, input_file)
            print(f"Successfully processed survey file: {input_file}")
            print("Input file updated in place")
            
    except Exception as e:
        logging.error(f"Error processing survey file: {e}")
        print(f"Error processing file: {e}")
        raise

def main():
    """Main entry point for command line usage."""
    import argparse
    
    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('survey_processing.log'),
            logging.StreamHandler()
        ]
    )
    
    parser = argparse.ArgumentParser(description='Process survey data files')
    parser.add_argument('input_file', help='Path to the input Excel file')
    parser.add_argument('-o', '--output', help='Path for the output file (optional)')
    
    args = parser.parse_args()
    
    process_survey_file(args.input_file, args.output)

if __name__ == "__main__":
    main()