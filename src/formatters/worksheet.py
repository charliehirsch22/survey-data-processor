"""
Worksheet formatting utilities.

This module provides functions for formatting Excel worksheets with headers, alignment, and structure.
"""

import logging
from typing import Dict

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string

from ..constants import (
    COL_WIDTH_NARROW,
    SHEET_DATA_MAP,
    SINGLE_SELECT_WITH_OTHER_WIDTHS,
    SINGLE_SELECT_WIDTHS,
)
from ..data_extractors.data_map_extractor import (
    find_question_text_from_data_map,
    find_column_l_text_from_data_map,
)
from .styles import create_thin_bottom_border


def apply_column_widths(worksheet: openpyxl.worksheet.worksheet.Worksheet, width_config: Dict[str, int]) -> None:
    """
    Apply column widths to a worksheet based on a configuration dictionary.

    Args:
        worksheet: The worksheet to apply widths to
        width_config: Dictionary mapping column letters to width values
    """
    for col, width in width_config.items():
        worksheet.column_dimensions[col].width = width


def setup_question_basic_formatting(question_ws: openpyxl.worksheet.worksheet.Worksheet, include_other: bool = False) -> None:
    """
    Apply basic formatting to a question worksheet: gridlines off, first 2 columns width 3.

    Args:
        question_ws: The worksheet to format
        include_other: Whether to include columns O, P, Q for "other specify" functionality
    """
    question_ws.sheet_view.showGridLines = False
    question_ws.column_dimensions['A'].width = COL_WIDTH_NARROW
    question_ws.column_dimensions['B'].width = COL_WIDTH_NARROW

    # Apply column widths based on question type
    if include_other:
        apply_column_widths(question_ws, SINGLE_SELECT_WITH_OTHER_WIDTHS)
    else:
        apply_column_widths(question_ws, SINGLE_SELECT_WIDTHS)


def add_question_text_and_section_header(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int, workbook: openpyxl.Workbook = None) -> None:
    """
    Add question text to C2 and section number to G4 from data map.

    Args:
        question_ws: The worksheet to add text to
        question_number: The question number (1-10)
        workbook: The workbook containing the data map tab (optional)
    """
    if workbook and SHEET_DATA_MAP in [ws.title for ws in workbook.worksheets]:
        data_map_ws = workbook[SHEET_DATA_MAP]
        question_text = find_question_text_from_data_map(data_map_ws, question_number)
        if question_text:
            question_ws['C2'] = question_text
            question_ws['C2'].font = Font(bold=True)
            logging.info(f"Added question text to C2: {question_text[:50]}...")
        else:
            question_ws['C2'] = f"Question {question_number} text not found"
            question_ws['C2'].font = Font(bold=True)
            logging.warning(f"Question text not found for question {question_number}")

        # Find and place column L text from data map in G4
        column_l_text = find_column_l_text_from_data_map(data_map_ws, question_number)
        if column_l_text:
            question_ws['G4'] = column_l_text
            logging.info(f"Added column L text to G4: {column_l_text}")
        else:
            question_ws['G4'] = f"Column L text not found for question {question_number}"
            logging.warning(f"Column L text not found for question {question_number}")
    else:
        question_ws['C2'] = f"Question {question_number} - data map not available"
        question_ws['C2'].font = Font(bold=True)
        question_ws['G4'] = "Data map not available"
        logging.warning("Data map not available for question text lookup")


def add_row4_headers(question_ws: openpyxl.worksheet.worksheet.Worksheet, include_q_header: bool = False) -> None:
    """
    Add standard headers to row 4 of a question worksheet.

    Args:
        question_ws: The worksheet to add headers to
        include_q_header: Whether to include Q4 header for "other specify" functionality
    """
    question_ws['C4'] = 'Response Text'
    question_ws['D4'] = 'N'
    question_ws['E4'] = '%'
    question_ws['I4'] = 'Filter Column #1'
    question_ws['J4'] = 'Filter #1'
    question_ws['K4'] = 'Filter Column #2'
    question_ws['L4'] = 'Filter #2'
    question_ws['M4'] = 'Filter Column #3'
    question_ws['N4'] = 'Filter #3'

    # Add thin bottom borders to header cells
    thin_bottom_border = create_thin_bottom_border()
    header_cells = ['C4', 'D4', 'E4', 'G4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4']
    if include_q_header:
        header_cells.append('Q4')

    for cell_ref in header_cells:
        question_ws[cell_ref].border = thin_bottom_border


def apply_center_alignment_to_columns(question_ws: openpyxl.worksheet.worksheet.Worksheet, include_q_column: bool = False) -> None:
    """
    Apply center alignment to standard columns in a question worksheet.

    Args:
        question_ws: The worksheet to apply alignment to
        include_q_column: Whether to include column Q for "other specify" functionality
    """
    center_alignment = Alignment(horizontal='center')
    center_columns = ['D', 'E', 'G', 'I', 'J', 'K', 'L', 'M', 'N']

    for col_letter in center_columns:
        for row in question_ws.iter_rows(min_col=column_index_from_string(col_letter),
                                       max_col=column_index_from_string(col_letter)):
            for cell in row:
                cell.alignment = center_alignment

    columns_str = "D:E, G, I:N"
    if include_q_column:
        columns_str = "D:E, G, I:Q"
    logging.info(f"Applied center alignment to columns {columns_str}")


def add_cross_cut_section(question_ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """
    Add the "Cross Cut" section below the response options.

    Args:
        question_ws: The worksheet to add the section to

    Returns:
        int: The row number where the Cross Cut section was added
    """
    # Find the last row with text in column C (should contain "<>")
    last_row_with_text = 6  # Start from row 6 where response options begin
    for row in range(6, question_ws.max_row + 1):
        cell_value = question_ws.cell(row=row, column=3).value  # Column C = 3
        if cell_value is not None and str(cell_value).strip():
            last_row_with_text = row

    # Place lowercase "x" in column B, two rows below the last row with text
    new_section_row = last_row_with_text + 2
    question_ws.cell(row=new_section_row, column=2, value='x')  # Column B = 2
    question_ws.cell(row=new_section_row, column=3, value='Cross Cut')  # Column C = 3

    # Add bottom border to cells C through N in the new section row
    thin_bottom_border = create_thin_bottom_border()
    for col in range(3, 15):  # Column C = 3, Column N = 14, so range(3, 15) covers C:N
        question_ws.cell(row=new_section_row, column=col).border = thin_bottom_border

    logging.info(f"Added 'x' marker in B{new_section_row} and 'Cross Cut' text in C{new_section_row} for additional analysis section")
    logging.info(f"Applied bottom border to cells C{new_section_row}:N{new_section_row}")

    return new_section_row
