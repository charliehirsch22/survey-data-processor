"""
Styling utilities for Excel worksheets.

This module provides functions to create reusable style objects for formatting Excel cells.
"""

from openpyxl.styles import Border, Side, PatternFill

from ..constants import COLOR_PALE_BLUE


def create_pale_blue_fill() -> PatternFill:
    """Create and return a pale blue fill pattern."""
    return PatternFill(
        start_color=COLOR_PALE_BLUE,
        end_color=COLOR_PALE_BLUE,
        fill_type='solid'
    )


def create_thin_border() -> Border:
    """Create and return a thin border on all sides."""
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def create_thin_bottom_border() -> Border:
    """Create and return a thin border on bottom only."""
    return Border(bottom=Side(style='thin'))
