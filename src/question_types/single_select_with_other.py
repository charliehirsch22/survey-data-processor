"""
Single select with "Other Specify" question type processor.

This module handles the setup and formatting of single select questions
that include "Other Specify" functionality.
"""

import logging
import re
from copy import copy

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

from ..constants import SHEET_DATA_MAP
from ..formatters.worksheet import (
    setup_question_basic_formatting,
    add_question_text_and_section_header,
    add_row4_headers,
    apply_center_alignment_to_columns,
    add_cross_cut_section,
)
from ..data_extractors.data_map_extractor import (
    extract_response_options,
    find_other_specify_child_text,
    extract_bracketed_text,
)


def cut_single_select_with_other(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int, workbook: openpyxl.Workbook = None) -> None:
    """
    Sets up a single select question with other specify functionality.

    This function is called when a question tab has the signature:
    "0, 0, 0, Simple Select, , 0, 0, 0, 0, Other Specify Parent"

    Args:
        question_ws: The question worksheet to set up
        question_number: The question number (1-10)
        workbook: The workbook containing the data map tab (optional)
    """
    try:
        logging.info(f"Setting up single select with other for question {question_number}")

        # Apply basic formatting with column widths
        setup_question_basic_formatting(question_ws, include_other=True)

        # Place lowercase x in cells B2 and P2
        question_ws['B2'] = 'x'
        question_ws['P2'] = 'x'

        # Add question text and section header
        add_question_text_and_section_header(question_ws, question_number, workbook)

        # Add headers to row 4
        add_row4_headers(question_ws, include_q_header=True)

        # Extract and place response options
        if workbook and SHEET_DATA_MAP in [ws.title for ws in workbook.worksheets]:
            data_map_ws = workbook[SHEET_DATA_MAP]
            extract_response_options(data_map_ws, question_ws, question_number)

            # Find and place "Other Specify Child" question text in Q2
            other_child_text = find_other_specify_child_text(data_map_ws, question_number)
            if other_child_text:
                question_ws['Q2'] = other_child_text
                question_ws['Q2'].font = Font(bold=True)
                logging.info(f"Added Other Specify Child text to Q2: {other_child_text[:50]}...")

                # Extract bracketed text from Q2 and place in Q4
                bracketed_text = extract_bracketed_text(other_child_text)
                if bracketed_text:
                    question_ws['Q4'] = bracketed_text
                    logging.info(f"Added bracketed text to Q4: {bracketed_text}")

                    # Add FILTER formula to Q6 with quote mark prefix to prevent calculation
                    filter_formula = '\'=FILTER(OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($Q$4, \'raw data\'!$C$2:$AJC$2, 0)-1), (OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($Q$4, \'raw data\'!$C$2:$AJC$2, 0)-1)<>"") * (IF($J$6="<>", TRUE, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($I$6, \'raw data\'!$C$2:$AJC$2, 0)-1)=$J$6)) * (IF($L$6="<>", TRUE, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($K$6, \'raw data\'!$C$2:$AJC$2, 0)-1)=$L$6)) * (IF($N$6="<>", TRUE, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($M$6, \'raw data\'!$C$2:$AJC$2, 0)-1)=$N$6)))'
                    question_ws['Q6'] = filter_formula
                    logging.info(f"Added FILTER formula to Q6")
                else:
                    logging.warning(f"No bracketed text found in Q2: {other_child_text}")
            else:
                logging.warning(f"No Other Specify Child text found for question {question_number}")

        # Apply center alignment to specified columns
        apply_center_alignment_to_columns(question_ws, include_q_column=False)

        # Add additional analysis section (Cross Cut)
        # Use cross_cut_row as the anchor for all Cross Cut section positioning
        cross_cut_row = add_cross_cut_section(question_ws)

        # Add filter labels using offsets from cross_cut_row
        question_ws.cell(row=cross_cut_row + 2, column=3, value='Filter Q #1')  # Column C = 3
        question_ws.cell(row=cross_cut_row + 3, column=3, value='Filter Column #1')
        question_ws.cell(row=cross_cut_row + 4, column=3, value='Filter #1')
        question_ws.cell(row=cross_cut_row + 5, column=3, value='Filter Q #2')
        question_ws.cell(row=cross_cut_row + 6, column=3, value='Filter Column #2')
        question_ws.cell(row=cross_cut_row + 7, column=3, value='Filter #2')
        logging.info(f"Added filter labels in column C from row {cross_cut_row + 2} to {cross_cut_row + 7}")

        # Add filter values in column D with blue font
        blue_font = Font(color="0000FF")
        # Row +2 and +5 left empty for Filter Q rows (formulas will be added later)
        question_ws.cell(row=cross_cut_row + 3, column=4, value='record').font = blue_font  # Column D = 4
        question_ws.cell(row=cross_cut_row + 4, column=4, value='<>').font = blue_font
        question_ws.cell(row=cross_cut_row + 6, column=4, value='record').font = blue_font
        question_ws.cell(row=cross_cut_row + 7, column=4, value='<>').font = blue_font
        logging.info(f"Added filter values (record/<>) in column D at rows {cross_cut_row + 3}, {cross_cut_row + 4}, {cross_cut_row + 6}, {cross_cut_row + 7} with blue font")

        # Add HumRead Filter labels in column C (same rows as OFFSET formulas in columns D-N)
        question_ws.cell(row=cross_cut_row + 9, column=3, value='HumRead Filter #1')
        question_ws.cell(row=cross_cut_row + 10, column=3, value='HumRead Filter #2')
        logging.info(f"Added HumRead Filter labels in column C at rows {cross_cut_row + 9} and {cross_cut_row + 10}")

        # Add blank row after HumRead Filter #2, then start data rows
        formula_row = cross_cut_row + 12  # Blank at +11, data starts at +12
        question_ws.cell(row=formula_row, column=3, value='=C6')  # Column C = 3
        logging.info(f"Added formula '=C6' in C{formula_row}")

        # Add OFFSET formulas for both header rows
        first_header_row = formula_row - 3  # Three rows above formula_row (at cross_cut_row + 9)
        second_header_row = formula_row - 2  # Two rows above formula_row (at cross_cut_row + 10)

        # Use flexible row references based on cross_cut_row offsets
        filter_q_1_row = cross_cut_row + 2  # Row with "Filter Q #1"
        filter_col_1_row = cross_cut_row + 3  # Row with "Filter Column #1"
        filter_1_row = cross_cut_row + 4  # Row with "Filter #1"
        filter_q_2_row = cross_cut_row + 5  # Row with "Filter Q #2"
        filter_col_2_row = cross_cut_row + 6  # Row with "Filter Column #2"
        filter_2_row = cross_cut_row + 7  # Row with "Filter #2"

        # Add Filter Q formulas to lookup English question text from data map
        filter_q_1_formula = f'=IF(D${filter_col_1_row}="record", "No filter", OFFSET(\'data map\'!$C$2, MATCH(D${filter_col_1_row}, \'data map\'!$L$2:$L$3200, 0)-1, 0))'
        question_ws.cell(row=filter_q_1_row, column=4, value=filter_q_1_formula)  # Column D = 4
        logging.info(f"Added Filter Q #1 OFFSET formula in D{filter_q_1_row} referencing D${filter_col_1_row}")

        filter_q_2_formula = f'=IF(D${filter_col_2_row}="record", "No filter", OFFSET(\'data map\'!$C$2, MATCH(D${filter_col_2_row}, \'data map\'!$L$2:$L$3200, 0)-1, 0))'
        question_ws.cell(row=filter_q_2_row, column=4, value=filter_q_2_formula)  # Column D = 4
        logging.info(f"Added Filter Q #2 OFFSET formula in D{filter_q_2_row} referencing D${filter_col_2_row}")

        # First header row - Filter #1
        first_header_formula = f'=IF(D${filter_1_row}="<>", "No filter", OFFSET(\'data map\'!$E$2, MATCH(D${filter_col_1_row}, \'data map\'!$L$2:$L$3200, 0)+D${filter_1_row},0))'
        question_ws.cell(row=first_header_row, column=4, value=first_header_formula)  # Column D = 4
        logging.info(f"Added first OFFSET formula in D{first_header_row} with flexible row references (D${filter_col_1_row} and D${filter_1_row})")

        # Second header row - Filter #2
        second_header_formula = f'=IF(D${filter_2_row}="<>", "No filter", OFFSET(\'data map\'!$E$2, MATCH(D${filter_col_2_row}, \'data map\'!$L$2:$L$3200, 0)+D${filter_2_row},0))'
        question_ws.cell(row=second_header_row, column=4, value=second_header_formula)  # Column D = 4
        logging.info(f"Added second OFFSET formula in D{second_header_row} with flexible row references (D${filter_col_2_row} and D${filter_2_row})")

        # Add COUNTIFS formula in the cell below (same row as =C6)
        countifs_formula = f"=COUNTIFS(OFFSET('raw data'!$C$3:$C$502, 0, MATCH($G$4, 'raw data'!$C$2:$AJC$2, 0)-1), $G6, OFFSET('raw data'!$C$3:$C$502, 0, MATCH(D${filter_col_1_row}, 'raw data'!$C$2:$AJC$2, 0)-1), D${filter_1_row}, OFFSET('raw data'!$C$3:$C$502, 0, MATCH(D${filter_col_2_row}, 'raw data'!$C$2:$AJC$2, 0)-1), D${filter_2_row})"
        question_ws.cell(row=formula_row, column=4, value=countifs_formula)  # Column D = 4
        logging.info(f"Added COUNTIFS formula in D{formula_row} with flexible row references")

        # Count the number of response options (from row 6 to the row with "<>" in column C)
        response_option_count = 0
        for row in range(6, question_ws.max_row + 1):
            cell_value = question_ws.cell(row=row, column=3).value  # Column C = 3
            if cell_value is not None and str(cell_value).strip():
                response_option_count += 1
                # Stop counting after we hit "<>"
                if str(cell_value).strip() == "<>":
                    break

        # Drag down the formula =C6 and COUNTIFS formula for the number of response options
        if response_option_count > 0:
            for i in range(response_option_count):
                current_formula_row = formula_row + i
                # Calculate the reference row (6 + i for C6, C7, C8, etc.)
                reference_row = 6 + i
                question_ws.cell(row=current_formula_row, column=3, value=f'=C{reference_row}')

                # Drag down COUNTIFS formula in column D, adjusting $G6 reference
                countifs_formula_dragged = f"=COUNTIFS(OFFSET('raw data'!$C$3:$C$502, 0, MATCH($G$4, 'raw data'!$C$2:$AJC$2, 0)-1), $G{reference_row}, OFFSET('raw data'!$C$3:$C$502, 0, MATCH(D${filter_col_1_row}, 'raw data'!$C$2:$AJC$2, 0)-1), D${filter_1_row}, OFFSET('raw data'!$C$3:$C$502, 0, MATCH(D${filter_col_2_row}, 'raw data'!$C$2:$AJC$2, 0)-1), D${filter_2_row})"
                question_ws.cell(row=current_formula_row, column=4, value=countifs_formula_dragged)

            logging.info(f"Dragged down formulas from row {formula_row} to {formula_row + response_option_count - 1} ({response_option_count} rows)")
            logging.info(f"Column C contains =C6 through =C{reference_row}, Column D contains COUNTIFS formulas")

            # Drag column D over to columns E through N for the entire new section
            # This includes Filter Column #1 through Filter #2 rows, OFFSET formula row, and all COUNTIFS formula rows
            start_drag_row = cross_cut_row + 2  # Start from "Filter Column #1" row
            end_drag_row = formula_row + response_option_count - 1  # End at the last response option row

            for row_num in range(start_drag_row, end_drag_row + 1):
                source_cell = question_ws.cell(row=row_num, column=4)  # Column D = 4
                source_value = source_cell.value

                # Copy from D to E through N (columns 5 through 14)
                for col_num in range(5, 15):  # E=5 through N=14
                    if source_value and isinstance(source_value, str) and source_value.startswith('='):
                        # This is a formula - adjust column references
                        adjusted_formula = source_value

                        # Calculate column offset (E is +1 from D, F is +2, etc.)
                        col_offset = col_num - 4  # D=4, so E=5 means offset of 1

                        # Replace column references like D$16, D$17, D$18, D$19 with adjusted columns
                        # But NOT absolute references like $G$4 (dollar before column)
                        # Pattern matches: letter (NOT preceded by $) + $ + digits (e.g., D$16 but not $G$4)
                        def adjust_col_ref(match):
                            col_letter = match.group(1)
                            dollar = match.group(2)
                            row_ref = match.group(3)

                            # Convert column letter to number, add offset, convert back
                            col_num_orig = column_index_from_string(col_letter)
                            new_col_num = col_num_orig + col_offset
                            new_col_letter = get_column_letter(new_col_num)

                            return f"{new_col_letter}{dollar}{row_ref}"

                        # Pattern to match column references like D$16 but NOT $G$4
                        # Negative lookbehind (?<!\$) ensures no $ before the column letter
                        adjusted_formula = re.sub(r'(?<!\$)([A-Z]+)(\$)(\d+)', adjust_col_ref, adjusted_formula)

                        target_cell = question_ws.cell(row=row_num, column=col_num)
                        target_cell.value = adjusted_formula
                        # Copy font from source cell
                        if source_cell.font:
                            target_cell.font = copy(source_cell.font)
                    else:
                        # Not a formula, just copy the value
                        target_cell = question_ws.cell(row=row_num, column=col_num)
                        target_cell.value = source_value
                        # Copy font from source cell
                        if source_cell.font:
                            target_cell.font = copy(source_cell.font)

            logging.info(f"Dragged column D (including filter values and formulas) over to columns E:N from row {start_drag_row} to {end_drag_row} with adjusted column references")

        # Ensure column widths and alignment are set correctly at the end (after all operations)
        question_ws.column_dimensions['D'].width = 16
        question_ws.column_dimensions['E'].width = 16
        question_ws.column_dimensions['F'].width = 16
        question_ws.column_dimensions['G'].width = 16
        question_ws.column_dimensions['H'].width = 16
        question_ws.column_dimensions['I'].width = 16
        question_ws.column_dimensions['J'].width = 16
        question_ws.column_dimensions['K'].width = 16
        question_ws.column_dimensions['L'].width = 16
        question_ws.column_dimensions['M'].width = 16
        question_ws.column_dimensions['N'].width = 16
        logging.info(f"Set column widths D:N (including F and H) to 16")

        # Apply center alignment to columns D:N for all rows
        center_alignment_final = openpyxl.styles.Alignment(horizontal='center')
        for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            for row in question_ws.iter_rows(min_col=openpyxl.utils.column_index_from_string(col_letter),
                                           max_col=openpyxl.utils.column_index_from_string(col_letter)):
                for cell in row:
                    cell.alignment = center_alignment_final
        logging.info(f"Applied center alignment to columns D:N (including F and H)")

        logging.info(f"Successfully set up single select with other for question {question_number}")

    except Exception as e:
        logging.error(f"Error setting up single select with other for question {question_number}: {e}")
        raise
