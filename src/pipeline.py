"""
Main processing pipeline orchestration.

This module coordinates the overall survey data processing workflow.
"""

import logging
from pathlib import Path

import openpyxl

from .constants import SHEET_DATA_MAP
from .setup.initial_setup import initial_set_up
from .setup.raw_data import raw_data_initial_setup
from .setup.data_map import data_map_initial_setup
from .setup.column_question_map import column_question_map_initial_setup
from .data_extractors.data_map_extractor import find_question_column_h_text
from .question_types.single_select_with_other import cut_single_select_with_other
from .question_types.single_select import cut_single_select
from .utils.excel_calculator import calculate_excel_formulas


def load_raw_excel_file(file_path: str) -> openpyxl.Workbook:
    """
    Loads an Excel file and returns an openpyxl Workbook object for processing.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        openpyxl.Workbook: The loaded workbook.

    Raises:
        FileNotFoundError: If the file doesn't exist.
        ValueError: If the file is not a valid Excel file.
    """
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if file_path_obj.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
            raise ValueError(f"Not a valid Excel file: {file_path}")

        logging.info(f"Loading Excel file: {file_path}")
        return openpyxl.load_workbook(file_path)
    except Exception as e:
        logging.error(f"Error loading Excel file {file_path}: {e}")
        raise


def save_processed_excel(workbook: openpyxl.Workbook, output_path: str) -> None:
    """
    Saves the processed workbook to the specified output path.

    Args:
        workbook (openpyxl.Workbook): The workbook to save.
        output_path (str): Path where to save the processed workbook.

    Raises:
        PermissionError: If unable to write to the file.
        OSError: If there's an OS-related error during saving.
    """
    try:
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)

        workbook.save(output_path)
        logging.info(f"Processed workbook saved successfully to: {output_path}")
    except Exception as e:
        logging.error(f"Error saving processed workbook to {output_path}: {e}")
        raise


def question_cutting_processor(workbook: openpyxl.Workbook) -> None:
    """
    Creates Q1-Q10 tabs for question processing.

    Args:
        workbook (openpyxl.Workbook): The workbook to add question tabs to.
    """
    try:
        logging.info("Starting question cutting processor...")

        # Create Q1-Q10 tabs
        create_question_tabs(workbook)

        logging.info("Question cutting processor completed successfully")

    except Exception as e:
        logging.error(f"Error during question cutting processor: {e}")
        raise


def create_question_tabs(workbook: openpyxl.Workbook) -> None:
    """
    Creates Q1-Q10 tabs for question processing.

    Args:
        workbook (openpyxl.Workbook): The workbook to add question tabs to.
    """
    try:
        logging.info("Creating Q1-Q10 tabs...")

        # Get the data map worksheet for column H lookups
        if SHEET_DATA_MAP not in [ws.title for ws in workbook.worksheets]:
            logging.warning(f"No '{SHEET_DATA_MAP}' tab found, creating tabs without column H data")
            data_map_ws = None
        else:
            data_map_ws = workbook[SHEET_DATA_MAP]

        # Create Q1 through Q10 tabs
        for i in range(1, 11):  # 1 to 10 inclusive
            tab_name = f"Q{i}"

            # Remove existing tab if it exists
            if tab_name in [ws.title for ws in workbook.worksheets]:
                workbook.remove(workbook[tab_name])
                logging.info(f"Removed existing {tab_name} tab")

            # Create new question tab
            question_ws = workbook.create_sheet(title=tab_name)

            # Add the question number in cell A1
            question_ws['A1'] = i

            # Find corresponding data from column H in data map
            if data_map_ws:
                column_h_text = find_question_column_h_text(data_map_ws, i)
                if column_h_text:
                    question_ws['A2'] = column_h_text
                    logging.info(f"Added column H text to {tab_name} A2: {column_h_text}")

                    # Detect question type and apply appropriate setup
                    if column_h_text == "0, 0, 0, Simple Select, , 0, 0, 0, 0, Other Specify Parent":
                        cut_single_select_with_other(question_ws, i, workbook)
                        logging.info(f"Applied single select with other setup to {tab_name}")
                    elif column_h_text == "0, 0, 0, Simple Select, , 0, 0, 0, 0, 0":
                        cut_single_select(question_ws, i, workbook)
                        logging.info(f"Applied single select setup to {tab_name}")

                else:
                    question_ws['A2'] = f"No data found for question {i}"
                    logging.info(f"No column H data found for question {i}")
            else:
                question_ws['A2'] = f"Data map not available"


            logging.info(f"Created {tab_name} tab")

        logging.info("Successfully created all Q1-Q10 tabs")

    except Exception as e:
        logging.error(f"Error creating question tabs: {e}")
        raise


def process_excel_file(input_path: str, output_path: str) -> None:
    """
    Main processing function that loads a raw Excel file and outputs a processed version.

    Args:
        input_path (str): Path to the input Excel file.
        output_path (str): Path for the output processed file.

    Raises:
        FileNotFoundError: If the input file doesn't exist.
        ValueError: If the input file is not a valid Excel file.
    """
    try:
        logging.info(f"Starting processing of {input_path}")

        # Load the raw Excel file
        workbook = load_raw_excel_file(input_path)

        # Perform initial setup
        initial_set_up(workbook)

        # Perform raw data initial setup
        raw_data_initial_setup(workbook)

        # Perform data map initial setup
        data_map_initial_setup(workbook)

        # Perform column question map initial setup
        column_question_map_initial_setup(workbook)

        # Save intermediate workbook with formulas (for debugging)
        base_path = Path(output_path)
        # Extract version number from output filename to keep consistent naming
        stem = base_path.stem
        if 'v' in stem:
            version_part = stem[stem.rfind('v'):]  # Gets 'v58' part
            formula_output_path = base_path.parent / f"test_processed_pilot{version_part}_formulas{base_path.suffix}"
        else:
            formula_output_path = base_path.parent / f"{stem}_formulas{base_path.suffix}"
        save_processed_excel(workbook, str(formula_output_path))
        logging.info(f"Saved intermediate workbook with formulas: {formula_output_path}")

        # Save workbook before Excel calculation
        save_processed_excel(workbook, output_path)

        # Calculate all formulas using Excel
        calculate_excel_formulas(output_path)

        # Reload workbook after Excel calculation with data_only=True to get calculated values
        workbook = openpyxl.load_workbook(output_path, data_only=True)

        # Perform question cutting
        question_cutting_processor(workbook)

        # Final save of the processed file
        save_processed_excel(workbook, output_path)

        logging.info(f"Successfully processed {input_path} -> {output_path}")

    except Exception as e:
        logging.error(f"Error processing Excel file {input_path}: {e}")
        raise
