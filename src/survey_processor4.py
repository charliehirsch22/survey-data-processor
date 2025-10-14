#!/usr/bin/env python3
"""
Survey Data Processor v4

Improved version of the survey processor with enhanced Excel file loading and output functionality.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Alignment


def load_raw_excel_file(file_path: str) -> openpyxl.Workbook:
    """
    Loads an Excel file and returns an openpyxl Workbook object for processing.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        openpyxl.Workbook: The loaded workbook.
    
    Raises:
        FileNotFoundError: If the file doesn't exist.
        ValueError: If the file is not a valid Excel file.
    """
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if file_path_obj.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
            raise ValueError(f"Not a valid Excel file: {file_path}")
        
        logging.info(f"Loading Excel file: {file_path}")
        return openpyxl.load_workbook(file_path)
    except Exception as e:
        logging.error(f"Error loading Excel file {file_path}: {e}")
        raise


def initial_set_up(workbook: openpyxl.Workbook) -> None:
    """
    Performs initial setup on the workbook:
    - Renames tabs "A", "B", or "raw data" to "raw data"
    - Renames tabs like "datamap" to "data map"
    - Adds blank tabs "column question map" and "loop variables"
    
    Args:
        workbook (openpyxl.Workbook): The workbook to set up.
    """
    try:
        logging.info("Starting initial setup...")
        
        # Get all worksheet names
        sheet_names = workbook.sheetnames
        logging.info(f"Found sheets: {sheet_names}")
        
        # Step 1: Rename tabs A, B, A1, or "raw data" to "raw data"
        for sheet_name in sheet_names:
            sheet_lower = sheet_name.lower()
            if (sheet_lower in ['a', 'b', 'raw data'] or 
                sheet_lower.startswith('a') and len(sheet_lower) <= 3):
                if sheet_lower != 'raw data':
                    workbook[sheet_name].title = 'raw data'
                    logging.info(f"Renamed '{sheet_name}' to 'raw data'")
                break
        
        # Step 2: Rename datamap-like tabs to "data map"
        for sheet_name in workbook.sheetnames:
            if 'datamap' in sheet_name.lower():
                workbook[sheet_name].title = 'data map'
                logging.info(f"Renamed '{sheet_name}' to 'data map'")
                break
        
        # Step 3: Add new blank tabs if they don't exist
        required_tabs = ['column question map', 'loop variables']
        existing_tabs = [ws.title.lower() for ws in workbook.worksheets]
        
        for tab_name in required_tabs:
            if tab_name.lower() not in existing_tabs:
                new_sheet = workbook.create_sheet(title=tab_name)
                logging.info(f"Created new tab: '{tab_name}'")
            else:
                logging.info(f"Tab '{tab_name}' already exists")
        
        logging.info("Initial setup completed successfully")
        
    except Exception as e:
        logging.error(f"Error during initial setup: {e}")
        raise


def raw_data_initial_setup(workbook: openpyxl.Workbook) -> None:
    """
    Performs initial setup on the raw data tab:
    - Inserts 2 columns at the front with width 3
    - Inserts 1 row at the top
    - Formats column headers (row 2) with borders and light blue background
    
    Args:
        workbook (openpyxl.Workbook): The workbook containing the raw data tab.
    """
    try:
        logging.info("Starting raw data initial setup...")
        
        # Get the raw data worksheet
        if 'raw data' not in [ws.title for ws in workbook.worksheets]:
            logging.warning("No 'raw data' tab found, skipping raw data setup")
            return
            
        ws = workbook['raw data']
        
        # Insert 2 columns at the front
        ws.insert_cols(1, 2)
        logging.info("Inserted 2 columns at the front")
        
        # Set width of the first 2 columns to 3
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 3
        logging.info("Set width of first 2 columns to 3")
        
        # Insert 1 row at the top
        ws.insert_rows(1, 1)
        logging.info("Inserted 1 row at the top")
        
        # Remove gridlines from the worksheet
        ws.sheet_view.showGridLines = False
        logging.info("Removed gridlines from worksheet")
        
        # Format column headers in row 2
        # Find the actual header row (should be row 2 after inserting 1 row)
        # But let's check both row 1 and 2 to be safe
        header_row = 2
        
        # Check if row 2 has headers, if not try row 1
        has_headers_row2 = any(ws.cell(row=2, column=col).value is not None for col in range(3, ws.max_column + 1))
        if not has_headers_row2:
            has_headers_row1 = any(ws.cell(row=1, column=col).value is not None for col in range(3, ws.max_column + 1))
            if has_headers_row1:
                header_row = 1
        
        logging.info(f"Using row {header_row} as header row")
        
        # Find the last column with data in the header row
        last_col = ws.max_column
        for col in range(3, ws.max_column + 1):  # Start from column 3 (C) since we inserted 2 columns
            if ws.cell(row=header_row, column=col).value is None:
                last_col = col - 1
                break
        
        # Create border style (outline on all sides)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create pale blue fill
        pale_blue_fill = PatternFill(
            start_color='E6F3FF',  # Pale blue color (lighter than ADD8E6)
            end_color='E6F3FF',
            fill_type='solid'
        )
        
        # Create center alignment
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply formatting to all header cells in the header row
        headers_formatted = 0
        for col in range(3, last_col + 1):  # Start from column 3 (C) since we inserted 2 columns
            cell = ws.cell(row=header_row, column=col)
            if cell.value is not None:  # Only format cells with content
                cell.border = thin_border
                cell.fill = pale_blue_fill
                cell.alignment = center_alignment
                headers_formatted += 1
        
        logging.info(f"Formatted {headers_formatted} column headers with borders, pale blue background, and center alignment")
        logging.info("Raw data initial setup completed successfully")
        
    except Exception as e:
        logging.error(f"Error during raw data initial setup: {e}")
        raise


def data_map_initial_setup(workbook: openpyxl.Workbook) -> None:
    """
    Performs initial setup on the data map tab:
    - Removes gridlines
    - Adds 2 columns on the left with width 3
    - Adds 3 rows at the top
    - Adds headers in row 2: "Question Info", "System Response Option", "Text Response Option", "Question Number"
    - Makes column F width 3
    
    Args:
        workbook (openpyxl.Workbook): The workbook containing the data map tab.
    """
    try:
        logging.info("Starting data map initial setup...")
        
        # Get the data map worksheet
        if 'data map' not in [ws.title for ws in workbook.worksheets]:
            logging.warning("No 'data map' tab found, skipping data map setup")
            return
            
        ws = workbook['data map']
        
        # Remove gridlines from the worksheet
        ws.sheet_view.showGridLines = False
        logging.info("Removed gridlines from data map worksheet")
        
        # Insert 2 columns at the front
        ws.insert_cols(1, 2)
        logging.info("Inserted 2 columns at the front")
        
        # Set width of the first 2 columns to 3
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 3
        logging.info("Set width of first 2 columns to 3")
        
        # Set width of column F to 3
        ws.column_dimensions['F'].width = 3
        logging.info("Set width of column F to 3")
        
        # Set column widths for data map layout
        ws.column_dimensions['C'].width = 50
        logging.info("Set width of column C to 50")
        
        # Set width of columns D and E to 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        logging.info("Set width of columns D:E to 13")
        
        # Set width of columns G through Z to 13
        for col_letter in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
            ws.column_dimensions[col_letter].width = 13
        logging.info("Set width of columns G:Z to 13")
        
        # Insert 3 rows at the top
        ws.insert_rows(1, 3)
        logging.info("Inserted 3 rows at the top")
        
        # Add headers in row 2
        ws['C2'] = "Question Info"
        ws['D2'] = "System Response Option"
        ws['E2'] = "Text Response Option"
        ws['G2'] = "Question Number"
        
        # Add additional headers from H to AG
        headers_h_to_ag = [
            "Question Type", "Question Text", "Question Code", "System or Survey Q", 
            "Question Prefix", "Section Marker", "Section Number", "Row Tag", 
            "Row Tag + Section", "Other Text Entry Flag", "Other Question Type Flag", 
            "Other Question Type", "System Question Flag", "System Question", 
            "Open Text", "Numerical", "Simple Select", "Multi Select (placeholder)", 
            "Rank", "Matrix", "Loop", "Double Loop", "Double Loop Flag #0", 
            "Double Loop Flag #1", "Double Loop Flag #2", "Double Loop Flag #3"
        ]
        
        # Start from column H (8) and add each header
        for i, header in enumerate(headers_h_to_ag):
            col_num = 8 + i  # H is column 8
            if col_num <= 33:  # AG is column 33
                ws.cell(row=2, column=col_num, value=header)
        
        logging.info("Added headers in row 2: Question Info, System Response Option, Text Response Option, Question Number, and additional headers H:AG")
        
        # Set text wrapping for column C and row 2
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        
        for cell in ws['C']:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        logging.info("Applied text wrapping to column C and row 2")
        
        # Format headers in row 2 with borders and pale blue background
        pale_blue_fill = PatternFill(
            start_color='E6F3FF',
            end_color='E6F3FF',
            fill_type='solid'
        )
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers_formatted = 0
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            if cell.value is not None:  # Only format cells with content
                cell.border = thin_border
                cell.fill = pale_blue_fill
                # Preserve the wrap_text alignment that was set earlier
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                headers_formatted += 1
        
        logging.info(f"Formatted {headers_formatted} column headers in row 2 with borders, pale blue background, and center alignment")
        
        # Auto-fit row height for row 2 to accommodate wrapped text
        # Calculate the maximum lines needed based on text content and column width
        max_lines = 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            if cell.value is not None:
                text = str(cell.value)
                # Get column width (default to 13 if not set)
                col_letter = get_column_letter(col)
                col_width = ws.column_dimensions[col_letter].width or 13
                
                # Estimate characters per line based on column width
                chars_per_line = max(1, int(col_width * 0.8))  # Rough estimate
                lines_needed = max(1, len(text) // chars_per_line + (1 if len(text) % chars_per_line > 0 else 0))
                
                # Account for word wrapping (longer words need more lines)
                words = text.split()
                if words:
                    longest_word = max(len(word) for word in words)
                    if longest_word > chars_per_line:
                        lines_needed = max(lines_needed, 2)
                
                max_lines = max(max_lines, lines_needed)
        
        # Set row height based on estimated lines (Excel default line height is about 15 points)
        estimated_height = max_lines * 15 + 5  # Add some padding
        ws.row_dimensions[2].height = estimated_height
        logging.info(f"Set row 2 height to {estimated_height} points to accommodate {max_lines} lines of wrapped text")
        
        # Add formulas starting in row 4
        # G4 formula
        ws['G4'] = "=INDEX('column question map'!$F$3:$F$939, MATCH('data map'!$L4, 'column question map'!$E$3:$E$939, 0))"
        # H4 formula
        ws['H4'] = '=U4&", "&V4&", "&W4&", "&X4&", "&Y4&", "&Z4&", "&AA4&", "&AB4&", "&AC4&", "&S4'
        # I4 formula
        ws['I4'] = '=INDEX($C$4:$C$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0))'
        # J4 formula
        ws['J4'] = '=SUBSTITUTE(SUBSTITUTE(LEFT(I4,FIND(":",I4)-1),"[",""),"]","")'
        # K4 formula
        ws['K4'] = '=IF(ISTEXT(LEFT(J4,1)),IF(EXACT(LEFT(J4,1),UPPER(LEFT(J4,1))),"Survey","System"),"First Char not Letter")'
        # L4 formula
        ws['L4'] = '=IF(K4="System","System",IF(ISNUMBER(FIND("_",J4)),LEFT(J4,FIND("_",J4)-1),IF(ISNUMBER(FIND("none",J4)),LEFT(J4,FIND("none",J4)-1),IF(ISNUMBER(FIND("r",J4)),LEFT(J4,FIND("r",J4)-1),J4))))'
        # M4 formula
        ws['M4'] = '=IF(AND(C4="", D4="", E4=""), "End", IF(AND(OFFSET(C4,-1,0)="", OFFSET(D4,-1,0)="", OFFSET(E4,-1,0)=""), "Start", "Mid"))'
        # N4 formula
        ws['N4'] = '=COUNTIFS($M$4:M4, "Start")'
        # O4 formula
        ws['O4'] = '=IF(M4="Start","Question Text",IF(O3="Question Text","Response Type",IF(ISNUMBER(D4),"Select Option",IF(AND(LEFT(D4,1)="[",RIGHT(D4,1)="]"),"Bracketed Sub-Question","End"))))'
        # P4 formula
        ws['P4'] = '=O4 & " " &N4'
        # Q4 formula
        ws['Q4'] = '=IF(ISNUMBER(SEARCH("oe]", C4)), "Other Text Entry", 0)'
        # R4 formula
        ws['R4'] = '=IFERROR(IF(Q4="Other Text Entry","Other Specify Child",IF(INDEX($Q$4:$Q$3200,MATCH(O4&" "&TEXT(N4+1,"0"),$P$4:$P$3200,0))="Other Text Entry","Other Specify Parent",0)),0)'
        # S4 formula
        ws['S4'] = '=INDEX($R$4:$R$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0))'
        # T4 formula
        ws['T4'] = '=IF(OR(ISNUMBER(SEARCH("[record]",C4)),ISNUMBER(SEARCH("[uuid]",C4)),ISNUMBER(SEARCH("[date]",C4)),ISNUMBER(SEARCH("[markers]",C4)),ISNUMBER(SEARCH("[status]",C4)),ISNUMBER(SEARCH("conditions: Conditions",C4)),ISNUMBER(SEARCH("[vlist]",C4)),ISNUMBER(SEARCH("[qtime]",C4)),ISNUMBER(SEARCH("[vos]",C4)),ISNUMBER(SEARCH("[vosr15oe]",C4)),ISNUMBER(SEARCH("[vbrowser]",C4)),ISNUMBER(SEARCH("[vbrowser15oe]",C4)),ISNUMBER(SEARCH("[vmobiledevice]",C4)),ISNUMBER(SEARCH("[vmobileos]",C4)),ISNUMBER(SEARCH("[start_date]",C4)),ISNUMBER(SEARCH("[vdropout]",C4)),ISNUMBER(SEARCH("[source]",C4)),ISNUMBER(SEARCH("[decLang]",C4)),ISNUMBER(SEARCH("[list]",C4)),ISNUMBER(SEARCH("[userAgent]",C4)),ISNUMBER(SEARCH("[fp_etag]",C4)),ISNUMBER(SEARCH("[fp_html5]",C4)),ISNUMBER(SEARCH("[fp_flash]",C4)),ISNUMBER(SEARCH("[fp_browser]",C4)),ISNUMBER(SEARCH("[dcua]",C4)),ISNUMBER(SEARCH("[url]",C4)),ISNUMBER(SEARCH("[session]",C4)),ISNUMBER(SEARCH("[s24627]",C4)),ISNUMBER(SEARCH("[s25023]",C4))),"System Question",0)'
        # U4 formula
        ws['U4'] = '=INDEX($T$4:$T$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0))'
        # V4 formula
        ws['V4'] = '=IF(INDEX($C$4:$C$3200, MATCH("Response Type "&N4, $P$4:$P$3200, 0)) = "Open text response", "Open Text", 0)'
        # W4 formula
        ws['W4'] = '=IF(INDEX($C$4:$C$3200, MATCH("Response Type "&N4, $P$4:$P$3200, 0)) = "Open numeric response", "Numerical", IF(AND(LEFT(INDEX($C$4:$C$3200, MATCH("Response Type "&N4, $P$4:$P$3200, 0)), 5) = "Value", ISERROR(MATCH("Select Option "&N4, $P$4:$P$3200, 0))), "Numerical", 0))'
        # X4 formula
        ws['X4'] = '=IF(AND(NOT(S4="Other Specify Child"), T4=0, U4=0, V4=0, W4=0, LEFT(INDEX($C$4:$C$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0)), 1) = "["), "Simple Select", 0)'
        # Y4 formula
        ws['Y4'] = '0'
        # Z4 formula
        ws['Z4'] = '=IF(ISNUMBER(SEARCH("rank", INDEX($C$4:$C$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0)))), "Rank", 0)'
        # AA4 formula
        ws['AA4'] = '=IF(AND(Z4=0, NOT((ISERROR(MATCH("Bracketed Sub-Question " & N4, $P$4:$P$3200, 0))))), "Matrix", 0)'
        # AB4 formula
        ws['AB4'] = '=IF(ISNUMBER(SEARCH("_", INDEX($C$4:$C$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0)))), "Loop", 0)'
        # AC4 formula
        ws['AC4'] = '=INDEX($AD$4:$AD$3200, MATCH("Question Text "&N4, $P$4:$P$3200, 0))'
        # AD4 formula
        ws['AD4'] = '=IF(AE4*AF4*AG4=1, "Double Loop", 0)'
        # AE4 formula
        ws['AE4'] = '=IFERROR(IF(AND(ISNUMBER(SEARCH("_",C4)),ISERROR(VALUE(MID(C4,SEARCH("_",C4)-1,1))),MID(C4,SEARCH("_",C4)-1,1)<>""),1,0), 0)'
        # AF4 formula
        ws['AF4'] = '=IF(AND(ISNUMBER(SEARCH("Lr",C4)),ISNUMBER(VALUE(MID(C4,SEARCH("Lr",C4)+2,1)))),1,0)'
        # AG4 formula
        ws['AG4'] = '=IF(OR(ISNUMBER(SEARCH("0r",C4)),ISNUMBER(SEARCH("1r",C4)),ISNUMBER(SEARCH("2r",C4)),ISNUMBER(SEARCH("3r",C4)),ISNUMBER(SEARCH("4r",C4)),ISNUMBER(SEARCH("5r",C4)),ISNUMBER(SEARCH("6r",C4)),ISNUMBER(SEARCH("7r",C4)),ISNUMBER(SEARCH("8r",C4)),ISNUMBER(SEARCH("9r",C4))),1,0)'
        logging.info("Added formulas to G4 through AG4 (27 total formulas)")
        
        # Find the last row with text in column C
        last_row_with_text = 1
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=3).value is not None and str(ws.cell(row=row, column=3).value).strip():
                last_row_with_text = row
        
        # Copy formulas down to one row below the last row with text in column C
        target_last_row = last_row_with_text + 1
        
        if target_last_row > 4:  # Only copy if there are rows to copy to
            # Copy formulas from G4:AG4 down to the target range (including G column)
            from openpyxl.utils import range_boundaries
            
            # Define the source range (G4:AG4) - including G column
            source_range = f"G4:AG4"
            # Define the target range (G5:AG{target_last_row}) - including G column
            target_range = f"G5:AG{target_last_row}"
            
            # Get source cells
            source_cells = ws[source_range]
            
            # Copy formulas to each row
            for row_idx, target_row in enumerate(range(5, target_last_row + 1)):
                for col_idx, source_cell in enumerate(source_cells[0]):  # source_cells[0] since it's a single row
                    target_cell = ws.cell(row=target_row, column=source_cell.column)
                    
                    if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                        # Copy the formula - openpyxl will automatically adjust relative references
                        target_cell.value = source_cell.value
                        
                        # Manually adjust relative row references in the formula
                        formula = source_cell.value
                        row_offset = target_row - 4
                        
                        # Replace relative row references (e.g., C4 -> C5, C6, etc.)
                        import re
                        def replace_row_ref(match):
                            col_ref = match.group(1)
                            row_ref = int(match.group(2))
                            # Only adjust relative references (not absolute ones with $)
                            if '$' not in match.group(0):
                                new_row = row_ref + row_offset
                                return f"{col_ref}{new_row}"
                            return match.group(0)
                        
                        # Pattern to match column+row references (e.g., C4, M4, etc.)
                        adjusted_formula = re.sub(r'([A-Z]+)(\d+)', replace_row_ref, formula)
                        target_cell.value = adjusted_formula
            
            copied_rows = target_last_row - 4
            logging.info(f"Copied formulas from G4:AG4 down to row {target_last_row} ({copied_rows} additional rows)")
        else:
            logging.info("No additional rows to copy formulas to")
        
        logging.info(f"Found last row with text in column C: row {last_row_with_text}")
        logging.info("Data map initial setup completed successfully")
        
    except Exception as e:
        logging.error(f"Error during data map initial setup: {e}")
        raise


def column_question_map_initial_setup(workbook: openpyxl.Workbook) -> None:
    """
    Performs initial setup on the column question map tab.
    
    Args:
        workbook (openpyxl.Workbook): The workbook containing the column question map tab.
    """
    try:
        logging.info("Starting column question map initial setup...")
        
        # Get the column question map worksheet
        if 'column question map' not in [ws.title for ws in workbook.worksheets]:
            logging.warning("No 'column question map' tab found, skipping column question map setup")
            return
            
        ws = workbook['column question map']
        
        # Remove gridlines from the worksheet
        ws.sheet_view.showGridLines = False
        logging.info("Removed gridlines from column question map worksheet")
        
        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 3
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        logging.info("Set column widths: A:B=3, C:H=20")
        
        # Add column headers in row 2, columns C through H
        headers = [
            "All question columns",
            "System or Survey", 
            "Question markers",
            "Question Number",
            "Unique question markers",
            "Question Number Map"
        ]
        
        for i, header in enumerate(headers):
            col = i + 3  # C=3, D=4, E=5, F=6, G=7, H=8
            ws.cell(row=2, column=col, value=header)
        
        logging.info("Added column headers in row 2: C2:H2")
        
        # Format headers in row 2 with borders and pale blue background
        pale_blue_fill = PatternFill(
            start_color='E6F3FF',
            end_color='E6F3FF',
            fill_type='solid'
        )
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        headers_formatted = 0
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            if cell.value is not None:  # Only format cells with content
                cell.border = thin_border
                cell.fill = pale_blue_fill
                cell.alignment = center_alignment
                headers_formatted += 1
        
        logging.info(f"Formatted {headers_formatted} column headers in row 2 with borders, pale blue background, and center alignment")
        
        # Copy column headers from raw data tab and transpose to column B
        if 'raw data' in [ws_temp.title for ws_temp in workbook.worksheets]:
            raw_data_ws = workbook['raw data']
            
            # Find the last column with text in row 2 of raw data
            last_col_with_text = 1
            for col in range(1, raw_data_ws.max_column + 1):
                if raw_data_ws.cell(row=2, column=col).value is not None and str(raw_data_ws.cell(row=2, column=col).value).strip():
                    last_col_with_text = col
            
            # Copy headers from C2 onwards in raw data tab
            headers_copied = 0
            current_row = 3  # Start at C3 in column question map
            
            for col in range(3, last_col_with_text + 1):  # Start from C2 (column 3)
                header_value = raw_data_ws.cell(row=2, column=col).value
                if header_value is not None:
                    # Paste as value (transposed) into column C of column question map
                    ws.cell(row=current_row, column=3, value=header_value)  # Column C = 3
                    headers_copied += 1
                    current_row += 1
            
            logging.info(f"Copied and transposed {headers_copied} column headers from raw data tab to column C (C3 onwards)")
        else:
            logging.warning("Raw data tab not found, skipping header copying")
        
        # Add formulas starting in row 3
        # D3 formula
        ws['D3'] = '=IF(ISTEXT(LEFT(C3,1)),IF(EXACT(LEFT(C3,1),UPPER(LEFT(C3,1))),"Survey","System"),"First Char not Letter")'
        # E3 formula
        ws['E3'] = '=IF(D3="System","System",IF(ISNUMBER(FIND("_",C3)),LEFT(C3,FIND("_",C3)-1),IF(ISNUMBER(FIND("none",C3)),LEFT(C3,FIND("none",C3)-1),IF(ISNUMBER(FIND("r",C3)),LEFT(C3,FIND("r",C3)-1),C3))))'
        # F3 formula
        ws['F3'] = '=IFERROR(INDEX($H$3:$H$200, MATCH($E3, $G$3:$G$200, 0)), "System")'
        logging.info("Added formulas to D3, E3, and F3")
        
        # Fill column H with sequential numbers 1-200 starting from H3
        for i in range(1, 201):  # 1 to 200
            row_num = i + 2  # H3 starts at row 3, so H3=1, H4=2, etc.
            ws.cell(row=row_num, column=8, value=i)  # Column H = 8
        
        logging.info("Added sequential numbers 1-200 in column H starting at H3")
        
        # Copy formulas D3:G3 down to the last row with text in column C
        # Find the last row with text in column C
        last_row_with_text = 1
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=3).value is not None and str(ws.cell(row=row, column=3).value).strip():
                last_row_with_text = row
        
        if last_row_with_text > 3:  # Only copy if there are rows below row 3
            # Copy formulas from D3:G3 down to the last row with text
            for row in range(4, last_row_with_text + 1):  # Start from row 4 (next row after formulas)
                for col in range(4, 8):  # D=4, E=5, F=6, G=7
                    source_cell = ws.cell(row=3, column=col)
                    target_cell = ws.cell(row=row, column=col)
                    
                    # Copy the formula, adjusting relative references
                    if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                        # Manually adjust relative row references in the formula
                        formula = source_cell.value
                        row_offset = row - 3
                        
                        # Replace relative row references (e.g., C3 -> C4, C5, etc.)
                        import re
                        def replace_row_ref(match):
                            col_ref = match.group(1)
                            row_ref = int(match.group(2))
                            # Only adjust relative references (not absolute ones with $)
                            if '$' not in match.group(0):
                                new_row = row_ref + row_offset
                                return f"{col_ref}{new_row}"
                            return match.group(0)
                        
                        # Pattern to match column+row references (e.g., C3, D3, etc.)
                        adjusted_formula = re.sub(r'([A-Z]+)(\d+)', replace_row_ref, formula)
                        target_cell.value = adjusted_formula
            
            copied_rows = last_row_with_text - 3
            logging.info(f"Copied formulas from D3:G3 down to row {last_row_with_text} ({copied_rows} additional rows)")
        else:
            logging.info("No additional rows to copy formulas to")
        
        logging.info(f"Found last row with text in column C: row {last_row_with_text}")
        
        # After formulas are copied, simulate the E column results to populate G with unique values
        # Since we can't evaluate Excel formulas in openpyxl, we'll simulate the E3 formula logic
        unique_values = []  # Use list to preserve order
        seen_values = set()  # Track what we've seen to avoid duplicates
        
        # Process each column header to simulate what column E formulas would produce
        for row in range(3, last_row_with_text + 1):
            # Get the column header from column C
            header_value = ws.cell(row=row, column=3).value  # Column C = 3
            if header_value is not None:
                header_str = str(header_value).strip()
                
                # Simulate the E3 formula logic: question marker extraction
                # Check if first character is a letter and uppercase (Survey vs System)
                if header_str and len(header_str) > 0:
                    first_char = header_str[0]
                    if first_char.isalpha() and first_char.isupper():
                        # This would be "Survey" - extract question marker
                        question_marker = header_str
                        
                        # Apply the same logic as E3 formula
                        if "_" in question_marker:
                            question_marker = question_marker[:question_marker.find("_")]
                        elif "none" in question_marker:
                            question_marker = question_marker[:question_marker.find("none")]
                        elif "r" in question_marker:
                            question_marker = question_marker[:question_marker.find("r")]
                        
                        # Add to unique list if not "System", not empty, and not already seen
                        if question_marker and question_marker != "System" and question_marker not in seen_values:
                            unique_values.append(question_marker)
                            seen_values.add(question_marker)
        
        # unique_values list now maintains the order from column E
        unique_list = unique_values
        
        # Populate column G starting at G3 with unique values
        for i, unique_value in enumerate(unique_list):
            row_num = i + 3  # Start at G3
            ws.cell(row=row_num, column=7, value=unique_value)  # Column G = 7
        
        logging.info(f"Populated column G with {len(unique_list)} unique question markers simulated from column C headers")
        
        # Apply center alignment to columns F, G, and H
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply center alignment to all of columns F, G, H (broader range to ensure coverage)
        for col in range(6, 9):  # Columns F=6, G=7, H=8
            for row in range(1, 1000):  # Apply to a large range to ensure all cells are covered
                ws.cell(row=row, column=col).alignment = center_alignment
        
        logging.info("Applied center alignment to entire columns F, G, and H")
        
        # Column G setup will be handled separately per user instructions
        
        logging.info("Column question map initial setup completed successfully")
        
    except Exception as e:
        logging.error(f"Error during column question map initial setup: {e}")
        raise



def calculate_excel_formulas(file_path: str) -> None:
    """
    Opens Excel to calculate all formulas in the workbook before proceeding with question cutting.
    This ensures that all formulas (especially column G lookups) are properly evaluated.
    
    Args:
        file_path (str): Path to the Excel file to calculate
    """
    try:
        import win32com.client
        logging.info("Opening Excel to calculate all formulas...")
        
        # Create Excel application
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False  # Keep Excel hidden
        excel_app.DisplayAlerts = False  # Disable alerts
        
        # Convert to absolute path for Excel
        import os
        abs_file_path = os.path.abspath(file_path)
        logging.info(f"Opening Excel file: {abs_file_path}")
        
        # Open the workbook
        workbook = excel_app.Workbooks.Open(abs_file_path)
        
        # Calculate all formulas
        excel_app.Calculate()
        logging.info("Calculated all formulas in Excel")
        
        # Save and close
        workbook.Save()
        workbook.Close()
        excel_app.Quit()
        
        logging.info("Excel calculation completed successfully")
        
    except ImportError:
        logging.warning("win32com.client not available - skipping Excel calculation. Install pywin32 for full functionality.")
    except Exception as e:
        logging.error(f"Error during Excel calculation: {e}")
        # Don't raise the error - continue with processing even if Excel calculation fails
        logging.warning("Continuing without Excel calculation - formulas may not be evaluated")


def question_cutting_processor(workbook: openpyxl.Workbook) -> None:
    """
    Creates Q1-Q10 tabs for question processing.
    
    Args:
        workbook (openpyxl.Workbook): The workbook to add question tabs to.
    """
    try:
        logging.info("Starting question cutting processor...")
        
        # Create Q1-Q10 tabs
        create_question_tabs(workbook)
        
        logging.info("Question cutting processor completed successfully")
        
    except Exception as e:
        logging.error(f"Error during question cutting processor: {e}")
        raise


def create_question_tabs(workbook: openpyxl.Workbook) -> None:
    """
    Creates Q1-Q10 tabs for question processing.
    
    Args:
        workbook (openpyxl.Workbook): The workbook to add question tabs to.
    """
    try:
        logging.info("Creating Q1-Q10 tabs...")
        
        # Get the data map worksheet for column H lookups
        if 'data map' not in [ws.title for ws in workbook.worksheets]:
            logging.warning("No 'data map' tab found, creating tabs without column H data")
            data_map_ws = None
        else:
            data_map_ws = workbook['data map']
        
        # Create Q1 through Q10 tabs
        for i in range(1, 11):  # 1 to 10 inclusive
            tab_name = f"Q{i}"
            
            # Remove existing tab if it exists
            if tab_name in [ws.title for ws in workbook.worksheets]:
                workbook.remove(workbook[tab_name])
                logging.info(f"Removed existing {tab_name} tab")
            
            # Create new question tab
            question_ws = workbook.create_sheet(title=tab_name)
            
            # Add the question number in cell A1
            question_ws['A1'] = i
            
            # Find corresponding data from column H in data map
            if data_map_ws:
                column_h_text = find_question_column_h_text(data_map_ws, i)
                if column_h_text:
                    question_ws['A2'] = column_h_text
                    logging.info(f"Added column H text to {tab_name} A2: {column_h_text}")
                    
                    # Detect question type and apply appropriate setup
                    if column_h_text == "0, 0, 0, Simple Select, , 0, 0, 0, 0, Other Specify Parent":
                        cut_single_select_with_other(question_ws, i, workbook)
                        logging.info(f"Applied single select with other setup to {tab_name}")
                    elif column_h_text == "0, 0, 0, Simple Select, , 0, 0, 0, 0, 0":
                        cut_single_select(question_ws, i, workbook)
                        logging.info(f"Applied single select setup to {tab_name}")
                    
                else:
                    question_ws['A2'] = f"No data found for question {i}"
                    logging.info(f"No column H data found for question {i}")
            else:
                question_ws['A2'] = f"Data map not available"
            
            
            logging.info(f"Created {tab_name} tab")
        
        logging.info("Successfully created all Q1-Q10 tabs")
        
    except Exception as e:
        logging.error(f"Error creating question tabs: {e}")
        raise


def find_question_column_h_text(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the first instance of a question number in column G and returns the corresponding column H value.
    Since column H contains formulas, we'll simulate the formula evaluation to get actual values.
    
    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)
        
    Returns:
        str: The evaluated value from column H, or None if not found
    """
    try:
        # Search through column G starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            cell_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7
            
            if cell_value is not None:
                cell_str = str(cell_value).strip()
                # Skip "System" entries
                if cell_str == "System":
                    continue
                
                # Check if this matches our question number
                if cell_str == str(question_number):
                    # Found the question, get column H value from same row
                    column_h_value = data_map_ws.cell(row=row, column=8).value  # Column H = 8
                    
                    # Check if it's a formula that needs evaluation
                    if column_h_value and isinstance(column_h_value, str) and column_h_value.startswith('='):
                        # Instead of trying to evaluate complex formulas, let's get the question info from column C
                        # which contains the actual question text and information
                        question_info = data_map_ws.cell(row=row, column=3).value  # Column C
                        if question_info and str(question_info).strip():
                            return str(question_info).strip()
                        else:
                            return f"Question {question_number} data from row {row}"
                    else:
                        # Not a formula, return the value directly
                        if column_h_value is not None:
                            return str(column_h_value).strip()
                        else:
                            return None
        
        # Question number not found
        return None
        
    except Exception as e:
        logging.error(f"Error finding column H text for question {question_number}: {e}")
        return None


def cut_single_select_with_other(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int, workbook: openpyxl.Workbook = None) -> None:
    """
    Sets up a single select question with other specify functionality.
    
    This function is called when a question tab has the signature:
    "0, 0, 0, Simple Select, , 0, 0, 0, 0, Other Specify Parent"
    
    Args:
        question_ws: The question worksheet to set up
        question_number: The question number (1-10)
        workbook: The workbook containing the data map tab (optional)
    """
    try:
        logging.info(f"Setting up single select with other for question {question_number}")
        
        # Apply basic formatting - ensure first 2 columns are blank with width 3 and remove gridlines
        question_ws.column_dimensions['A'].width = 3
        question_ws.column_dimensions['B'].width = 3
        question_ws.sheet_view.showGridLines = False
        
        # Set column widths as specified
        question_ws.column_dimensions['C'].width = 20
        question_ws.column_dimensions['D'].width = 16
        question_ws.column_dimensions['E'].width = 16
        question_ws.column_dimensions['F'].width = 3
        question_ws.column_dimensions['G'].width = 16
        question_ws.column_dimensions['H'].width = 3
        question_ws.column_dimensions['I'].width = 16
        question_ws.column_dimensions['J'].width = 16
        question_ws.column_dimensions['K'].width = 16
        question_ws.column_dimensions['L'].width = 16
        question_ws.column_dimensions['M'].width = 16
        question_ws.column_dimensions['N'].width = 16
        question_ws.column_dimensions['O'].width = 3
        question_ws.column_dimensions['P'].width = 3
        question_ws.column_dimensions['Q'].width = 13
        
        # Place lowercase x in cells B2 and P2
        question_ws['B2'] = 'x'
        question_ws['P2'] = 'x'
        
        # Find and place question text from data map column C
        if workbook and 'data map' in [ws.title for ws in workbook.worksheets]:
            data_map_ws = workbook['data map']
            question_text = find_question_text_from_data_map(data_map_ws, question_number)
            if question_text:
                question_ws['C2'] = question_text
                # Make C2 bold
                question_ws['C2'].font = openpyxl.styles.Font(bold=True)
                logging.info(f"Added question text to C2: {question_text[:50]}...")
            else:
                question_ws['C2'] = f"Question {question_number} text not found"
                question_ws['C2'].font = openpyxl.styles.Font(bold=True)
                logging.warning(f"Question text not found for question {question_number}")
                
            # Find and place column L text from data map in G4
            column_l_text = find_column_l_text_from_data_map(data_map_ws, question_number)
            if column_l_text:
                question_ws['G4'] = column_l_text
                logging.info(f"Added column L text to G4: {column_l_text}")
            else:
                question_ws['G4'] = f"Column L text not found for question {question_number}"
                logging.warning(f"Column L text not found for question {question_number}")
                
        else:
            question_ws['C2'] = f"Question {question_number} - data map not available"
            question_ws['C2'].font = openpyxl.styles.Font(bold=True)
            question_ws['G4'] = "Data map not available"
            logging.warning("Data map not available for question text lookup")
        
        # Add headers to row 4
        question_ws['C4'] = 'Response Text'
        question_ws['D4'] = 'N'
        question_ws['E4'] = '%'
        question_ws['I4'] = 'Filter Column #1'
        question_ws['J4'] = 'Filter #1'
        question_ws['K4'] = 'Filter Column #2'
        question_ws['L4'] = 'Filter #2'
        question_ws['M4'] = 'Filter Column #3'
        question_ws['N4'] = 'Filter #3'
        
        # Add thin bottom borders to header cells
        thin_bottom_border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        header_cells = ['C4', 'D4', 'E4', 'G4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'Q4']
        for cell_ref in header_cells:
            question_ws[cell_ref].border = thin_bottom_border
        
        # Extract and place response options
        if workbook and 'data map' in [ws.title for ws in workbook.worksheets]:
            data_map_ws = workbook['data map']
            extract_response_options(data_map_ws, question_ws, question_number)
            
            # Find and place "Other Specify Child" question text in Q2
            other_child_text = find_other_specify_child_text(data_map_ws, question_number)
            if other_child_text:
                question_ws['Q2'] = other_child_text
                question_ws['Q2'].font = openpyxl.styles.Font(bold=True)
                logging.info(f"Added Other Specify Child text to Q2: {other_child_text[:50]}...")
                
                # Extract bracketed text from Q2 and place in Q4
                bracketed_text = extract_bracketed_text(other_child_text)
                if bracketed_text:
                    question_ws['Q4'] = bracketed_text
                    logging.info(f"Added bracketed text to Q4: {bracketed_text}")
                    
                    # Add FILTER formula to Q6 with quote mark prefix to prevent calculation
                    filter_formula = '\'=FILTER(OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($Q$4, \'raw data\'!$C$2:$AJC$2, 0)-1), (OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($Q$4, \'raw data\'!$C$2:$AJC$2, 0)-1)<>"") * (IF($J$6="<>", TRUE, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($I$6, \'raw data\'!$C$2:$AJC$2, 0)-1)=$J$6)) * (IF($L$6="<>", TRUE, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($K$6, \'raw data\'!$C$2:$AJC$2, 0)-1)=$L$6)) * (IF($N$6="<>", TRUE, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($M$6, \'raw data\'!$C$2:$AJC$2, 0)-1)=$N$6)))'
                    question_ws['Q6'] = filter_formula
                    logging.info(f"Added FILTER formula to Q6")
                else:
                    logging.warning(f"No bracketed text found in Q2: {other_child_text}")
            else:
                logging.warning(f"No Other Specify Child text found for question {question_number}")
        
        # Apply center alignment to specified columns
        center_alignment = openpyxl.styles.Alignment(horizontal='center')
        center_columns = ['D', 'E', 'G', 'I', 'J', 'K', 'L', 'M', 'N']

        for col_letter in center_columns:
            for row in question_ws.iter_rows(min_col=openpyxl.utils.column_index_from_string(col_letter),
                                           max_col=openpyxl.utils.column_index_from_string(col_letter)):
                for cell in row:
                    cell.alignment = center_alignment

        logging.info(f"Applied center alignment to columns D:E, G, I:N")

        # Add additional analysis section
        # Find the last row with text in column C (should contain "<>")
        last_row_with_text = 6  # Start from row 6 where response options begin
        for row in range(6, question_ws.max_row + 1):
            cell_value = question_ws.cell(row=row, column=3).value  # Column C = 3
            if cell_value is not None and str(cell_value).strip():
                last_row_with_text = row

        # Place lowercase "x" in column B, two rows below the last row with text
        new_section_row = last_row_with_text + 2
        question_ws.cell(row=new_section_row, column=2, value='x')  # Column B = 2
        question_ws.cell(row=new_section_row, column=3, value='Cross Cut')  # Column C = 3

        # Add bottom border to cells C through N in the new section row
        thin_bottom_border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        for col in range(3, 15):  # Column C = 3, Column N = 14, so range(3, 15) covers C:N
            question_ws.cell(row=new_section_row, column=col).border = thin_bottom_border

        logging.info(f"Added 'x' marker in B{new_section_row} and 'Cross Cut' text in C{new_section_row} for additional analysis section")
        logging.info(f"Applied bottom border to cells C{new_section_row}:N{new_section_row}")

        # Add filter labels starting two rows below "Cross Cut"
        filter_start_row = new_section_row + 2
        question_ws.cell(row=filter_start_row, column=3, value='Filter Column #1')  # Column C = 3
        question_ws.cell(row=filter_start_row + 1, column=3, value='Filter #1')
        question_ws.cell(row=filter_start_row + 2, column=3, value='Filter Column #2')
        question_ws.cell(row=filter_start_row + 3, column=3, value='Filter #2')
        logging.info(f"Added filter labels in column C from row {filter_start_row} to {filter_start_row + 3}")

        # Add filter values in column D
        question_ws.cell(row=filter_start_row, column=4, value='record')  # Column D = 4
        question_ws.cell(row=filter_start_row + 1, column=4, value='<>')
        question_ws.cell(row=filter_start_row + 2, column=4, value='record')
        question_ws.cell(row=filter_start_row + 3, column=4, value='<>')
        logging.info(f"Added filter values (record/<>) in column D from row {filter_start_row} to {filter_start_row + 3}")

        # Add formula 3 rows below "Filter #2" in column C
        formula_row = filter_start_row + 3 + 3  # "Filter #2" is at filter_start_row + 3, then add 3 more rows
        question_ws.cell(row=formula_row, column=3, value='=C6')  # Column C = 3
        logging.info(f"Added formula '=C6' in C{formula_row}")

        # Add OFFSET formula one row up and one column to the right (column D)
        offset_formula_row = formula_row - 1
        # Use flexible row references based on where "Filter Column #1" and "Filter #1" are located
        filter_col_1_row = filter_start_row  # Row with "Filter Column #1"
        filter_1_row = filter_start_row + 1  # Row with "Filter #1"
        offset_formula = f"=OFFSET('data map'!$E$2, MATCH(D${filter_col_1_row}, 'data map'!$L$2:$L$3200, 0)+D${filter_1_row},0)"
        question_ws.cell(row=offset_formula_row, column=4, value=offset_formula)  # Column D = 4
        logging.info(f"Added OFFSET formula in D{offset_formula_row} with flexible row references (D${filter_col_1_row} and D${filter_1_row})")

        # Add COUNTIFS formula in the cell below (same row as =C6)
        # Use flexible row references based on where filter rows are located
        filter_col_2_row = filter_start_row + 2  # Row with "Filter Column #2"
        filter_2_row = filter_start_row + 3  # Row with "Filter #2"
        countifs_formula = f"=COUNTIFS(OFFSET('raw data'!$C$3:$C$502, 0, MATCH($G$4, 'raw data'!$C$2:$AJC$2, 0)-1), $G6, OFFSET('raw data'!$C$3:$C$502, 0, MATCH(D${filter_col_1_row}, 'raw data'!$C$2:$AJC$2, 0)-1), D${filter_1_row}, OFFSET('raw data'!$C$3:$C$502, 0, MATCH(D${filter_col_2_row}, 'raw data'!$C$2:$AJC$2, 0)-1), D${filter_2_row})"
        question_ws.cell(row=formula_row, column=4, value=countifs_formula)  # Column D = 4
        logging.info(f"Added COUNTIFS formula in D{formula_row} with flexible row references")

        # Count the number of response options (from row 6 to the row with "<>" in column C)
        response_option_count = 0
        for row in range(6, question_ws.max_row + 1):
            cell_value = question_ws.cell(row=row, column=3).value  # Column C = 3
            if cell_value is not None and str(cell_value).strip():
                response_option_count += 1
                # Stop counting after we hit "<>"
                if str(cell_value).strip() == "<>":
                    break

        # Drag down the formula =C6 and COUNTIFS formula for the number of response options
        if response_option_count > 0:
            for i in range(response_option_count):
                current_formula_row = formula_row + i
                # Calculate the reference row (6 + i for C6, C7, C8, etc.)
                reference_row = 6 + i
                question_ws.cell(row=current_formula_row, column=3, value=f'=C{reference_row}')

                # Drag down COUNTIFS formula in column D, adjusting $G6 reference
                countifs_formula_dragged = f"=COUNTIFS(OFFSET('raw data'!$C$3:$C$502, 0, MATCH($G$4, 'raw data'!$C$2:$AJC$2, 0)-1), $G{reference_row}, OFFSET('raw data'!$C$3:$C$502, 0, MATCH(D$16, 'raw data'!$C$2:$AJC$2, 0)-1), D$17, OFFSET('raw data'!$C$3:$C$502, 0, MATCH(D$18, 'raw data'!$C$2:$AJC$2, 0)-1), D$19)"
                question_ws.cell(row=current_formula_row, column=4, value=countifs_formula_dragged)

            logging.info(f"Dragged down formulas from row {formula_row} to {formula_row + response_option_count - 1} ({response_option_count} rows)")
            logging.info(f"Column C contains =C6 through =C{reference_row}, Column D contains COUNTIFS formulas")

            # Drag column D over to columns E through N for the entire new section
            # This includes Filter Column #1 through Filter #2 rows, OFFSET formula row, and all COUNTIFS formula rows
            start_drag_row = filter_start_row  # Start from "Filter Column #1" row
            end_drag_row = formula_row + response_option_count - 1  # End at the last response option row

            for row_num in range(start_drag_row, end_drag_row + 1):
                source_cell = question_ws.cell(row=row_num, column=4)  # Column D = 4
                source_value = source_cell.value

                # Copy from D to E through N (columns 5 through 14)
                for col_num in range(5, 15):  # E=5 through N=14
                    if source_value and isinstance(source_value, str) and source_value.startswith('='):
                        # This is a formula - adjust column references
                        import re
                        adjusted_formula = source_value

                        # Calculate column offset (E is +1 from D, F is +2, etc.)
                        col_offset = col_num - 4  # D=4, so E=5 means offset of 1

                        # Replace column references like D$16, D$17, D$18, D$19 with adjusted columns
                        # But NOT absolute references like $G$4 (dollar before column)
                        # Pattern matches: letter (NOT preceded by $) + $ + digits (e.g., D$16 but not $G$4)
                        def adjust_col_ref(match):
                            col_letter = match.group(1)
                            dollar = match.group(2)
                            row_ref = match.group(3)

                            # Convert column letter to number, add offset, convert back
                            from openpyxl.utils import column_index_from_string, get_column_letter
                            col_num_orig = column_index_from_string(col_letter)
                            new_col_num = col_num_orig + col_offset
                            new_col_letter = get_column_letter(new_col_num)

                            return f"{new_col_letter}{dollar}{row_ref}"

                        # Pattern to match column references like D$16 but NOT $G$4
                        # Negative lookbehind (?<!\$) ensures no $ before the column letter
                        adjusted_formula = re.sub(r'(?<!\$)([A-Z]+)(\$)(\d+)', adjust_col_ref, adjusted_formula)

                        question_ws.cell(row=row_num, column=col_num, value=adjusted_formula)
                    else:
                        # Not a formula, just copy the value
                        question_ws.cell(row=row_num, column=col_num, value=source_value)

            logging.info(f"Dragged column D (including filter values and formulas) over to columns E:N from row {start_drag_row} to {end_drag_row} with adjusted column references")

        # Ensure column widths and alignment are set correctly at the end (after all operations)
        question_ws.column_dimensions['D'].width = 16
        question_ws.column_dimensions['E'].width = 16
        question_ws.column_dimensions['F'].width = 16
        question_ws.column_dimensions['G'].width = 16
        question_ws.column_dimensions['H'].width = 16
        question_ws.column_dimensions['I'].width = 16
        question_ws.column_dimensions['J'].width = 16
        question_ws.column_dimensions['K'].width = 16
        question_ws.column_dimensions['L'].width = 16
        question_ws.column_dimensions['M'].width = 16
        question_ws.column_dimensions['N'].width = 16
        logging.info(f"Set column widths D:N (including F and H) to 16")

        # Apply center alignment to columns D:N for all rows
        center_alignment_final = openpyxl.styles.Alignment(horizontal='center')
        for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            for row in question_ws.iter_rows(min_col=openpyxl.utils.column_index_from_string(col_letter),
                                           max_col=openpyxl.utils.column_index_from_string(col_letter)):
                for cell in row:
                    cell.alignment = center_alignment_final
        logging.info(f"Applied center alignment to columns D:N (including F and H)")

        logging.info(f"Successfully set up single select with other for question {question_number}")

    except Exception as e:
        logging.error(f"Error setting up single select with other for question {question_number}: {e}")
        raise


def cut_single_select(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int, workbook: openpyxl.Workbook = None) -> None:
    """
    Sets up a single select question worksheet with formatting, response options, and formulas.
    Similar to cut_single_select_with_other but excludes columns O, P, Q functionality.
    
    Args:
        question_ws: The worksheet to set up
        question_number: The question number (1-10)
        workbook: The parent workbook containing data map for response option extraction
    """
    try:
        logging.info(f"Setting up single select for question {question_number}")
        
        # Apply basic formatting - ensure first 2 columns are blank with width 3 and remove gridlines
        question_ws.column_dimensions['A'].width = 3
        question_ws.column_dimensions['B'].width = 3
        question_ws.sheet_view.showGridLines = False
        
        # Set column widths as specified (excluding O, P, Q)
        question_ws.column_dimensions['C'].width = 20
        question_ws.column_dimensions['D'].width = 13
        question_ws.column_dimensions['E'].width = 13
        question_ws.column_dimensions['F'].width = 3
        question_ws.column_dimensions['G'].width = 13
        question_ws.column_dimensions['H'].width = 3
        question_ws.column_dimensions['I'].width = 13
        question_ws.column_dimensions['J'].width = 13
        question_ws.column_dimensions['K'].width = 13
        question_ws.column_dimensions['L'].width = 13
        question_ws.column_dimensions['M'].width = 13
        question_ws.column_dimensions['N'].width = 13
        
        # Place lowercase x in cell B2 only (no P2 for single select)
        question_ws['B2'] = 'x'
        
        # Find and place question text from data map column C
        if workbook and 'data map' in [ws.title for ws in workbook.worksheets]:
            data_map_ws = workbook['data map']
            question_text = find_question_text_from_data_map(data_map_ws, question_number)
            if question_text:
                question_ws['C2'] = question_text
                # Make C2 bold
                question_ws['C2'].font = openpyxl.styles.Font(bold=True)
                logging.info(f"Added question text to C2: {question_text[:50]}...")
            else:
                question_ws['C2'] = f"Question {question_number} text not found"
                question_ws['C2'].font = openpyxl.styles.Font(bold=True)
                logging.warning(f"Question text not found for question {question_number}")
                
            # Find and place column L text from data map in G4
            column_l_text = find_column_l_text_from_data_map(data_map_ws, question_number)
            if column_l_text:
                question_ws['G4'] = column_l_text
                logging.info(f"Added column L text to G4: {column_l_text}")
            else:
                question_ws['G4'] = f"Column L text not found for question {question_number}"
                logging.warning(f"Column L text not found for question {question_number}")
                
        else:
            question_ws['C2'] = f"Question {question_number} - data map not available"
            question_ws['C2'].font = openpyxl.styles.Font(bold=True)
            question_ws['G4'] = "Data map not available"
            logging.warning("Data map not available for question text lookup")
        
        # Add headers to row 4 (excluding Q4)
        question_ws['C4'] = 'Response Text'
        question_ws['D4'] = 'N'
        question_ws['E4'] = '%'
        question_ws['I4'] = 'Filter Column #1'
        question_ws['J4'] = 'Filter #1'
        question_ws['K4'] = 'Filter Column #2'
        question_ws['L4'] = 'Filter #2'
        question_ws['M4'] = 'Filter Column #3'
        question_ws['N4'] = 'Filter #3'
        
        # Add thin bottom borders to header cells (excluding Q4)
        thin_bottom_border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
        header_cells = ['C4', 'D4', 'E4', 'G4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4']
        for cell_ref in header_cells:
            question_ws[cell_ref].border = thin_bottom_border
        
        # Extract and place response options (no Other Specify Child functionality)
        if workbook and 'data map' in [ws.title for ws in workbook.worksheets]:
            data_map_ws = workbook['data map']
            extract_response_options(data_map_ws, question_ws, question_number)
        
        # Apply center alignment to specified columns (excluding Q)
        center_alignment = openpyxl.styles.Alignment(horizontal='center')
        center_columns = ['D', 'E', 'G', 'I', 'J', 'K', 'L', 'M', 'N']
        
        for col_letter in center_columns:
            for row in question_ws.iter_rows(min_col=openpyxl.utils.column_index_from_string(col_letter), 
                                           max_col=openpyxl.utils.column_index_from_string(col_letter)):
                for cell in row:
                    cell.alignment = center_alignment
        
        logging.info(f"Applied center alignment to columns D:E, G, I:N")
        logging.info(f"Successfully set up single select for question {question_number}")
        
    except Exception as e:
        logging.error(f"Error setting up single select for question {question_number}: {e}")
        raise


def find_question_text_from_data_map(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the question text from column C of the data map in the same row as the first instance
    of the question number in column G.
    
    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)
        
    Returns:
        str: The question text from column C, or None if not found
    """
    try:
        # Search through column G starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            cell_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7
            
            if cell_value is not None:
                cell_str = str(cell_value).strip()
                
                # Check if this matches our question number
                if cell_str == str(question_number):
                    # Found the question, get column C value from same row
                    question_text = data_map_ws.cell(row=row, column=3).value  # Column C = 3
                    if question_text:
                        return str(question_text).strip()
                    else:
                        return None
        
        return None
        
    except Exception as e:
        logging.error(f"Error finding question text for question {question_number}: {e}")
        return None


def find_column_l_text_from_data_map(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the text from column L of the data map in the same row as the first instance
    of the question number in column G.
    
    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)
        
    Returns:
        str: The text from column L, or None if not found
    """
    try:
        # Search through column G starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            cell_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7
            
            if cell_value is not None:
                cell_str = str(cell_value).strip()
                
                # Check if this matches our question number
                if cell_str == str(question_number):
                    # Found the question, get column L value from same row
                    column_l_text = data_map_ws.cell(row=row, column=12).value  # Column L = 12
                    if column_l_text:
                        return str(column_l_text).strip()
                    else:
                        return None
        
        return None
        
    except Exception as e:
        logging.error(f"Error finding column L text for question {question_number}: {e}")
        return None


def extract_response_options(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> None:
    """
    Extracts response options from the data map and places them in the question worksheet.
    
    Args:
        data_map_ws: The data map worksheet
        question_ws: The question worksheet to populate
        question_number: The question number to process
    """
    try:
        logging.info(f"Extracting response options for question {question_number}")
        
        # Step 1: Find the section number from column N
        section_number = find_section_number_from_data_map(data_map_ws, question_number)
        if not section_number:
            logging.warning(f"No section number found for question {question_number}")
            return
        
        logging.info(f"Found section number for question {question_number}: {section_number}")
        
        # Step 2: Find all rows with "Select Option [section]" in column P
        target_pattern = f"Select Option {section_number}"
        response_rows = []
        
        for row in range(4, data_map_ws.max_row + 1):
            column_p_value = data_map_ws.cell(row=row, column=16).value  # Column P = 16
            if column_p_value and str(column_p_value).strip() == target_pattern:
                response_rows.append(row)
        
        logging.info(f"Found {len(response_rows)} response option rows for '{target_pattern}'")
        
        if not response_rows:
            logging.warning(f"No response options found for pattern '{target_pattern}'")
            return
        
        # Step 3: Extract response numbers (column D) and text (column E)
        current_row = 6  # Start placing at row 6
        
        for data_row in response_rows:
            # Get response number from column D
            response_number = data_map_ws.cell(row=data_row, column=4).value  # Column D = 4
            # Get response text from column E  
            response_text = data_map_ws.cell(row=data_row, column=5).value  # Column E = 5
            
            # Place in question worksheet
            if response_number is not None:
                question_ws.cell(row=current_row, column=7, value=response_number)  # Column G = 7
            if response_text is not None:
                question_ws.cell(row=current_row, column=3, value=response_text)  # Column C = 3
            
            logging.info(f"Added response option {current_row-5}: {response_number} - {response_text}")
            current_row += 1
        
        # Step 4: Add "<>" terminator in the next row
        question_ws.cell(row=current_row, column=7, value="<>")  # Column G
        question_ws.cell(row=current_row, column=3, value="<>")  # Column C
        
        logging.info(f"Added terminator '<>' at row {current_row}")
        
        # Step 5: Add COUNTIFS formula to column D starting at row 6
        formula = '=COUNTIFS(OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($G$4, \'raw data\'!$C$2:$AJC$2, 0)-1), $G6, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($I6, \'raw data\'!$C$2:$AJC$2, 0)-1), $J6, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($K6, \'raw data\'!$C$2:$AJC$2, 0)-1), $L6, OFFSET(\'raw data\'!$C$3:$C$502, 0, MATCH($M6, \'raw data\'!$C$2:$AJC$2, 0)-1), $N6)'
        
        # Add formula to D6 and drag down to include the terminator row (current_row has "<>")
        for formula_row in range(6, current_row + 1):  # Include the "<>" row
            # Adjust the formula for the current row by replacing "6" with the current row number
            adjusted_formula = formula.replace('$G6', f'$G{formula_row}').replace('$I6', f'$I{formula_row}').replace('$J6', f'$J{formula_row}').replace('$K6', f'$K{formula_row}').replace('$L6', f'$L{formula_row}').replace('$M6', f'$M{formula_row}').replace('$N6', f'$N{formula_row}')
            question_ws.cell(row=formula_row, column=4, value=adjusted_formula)  # Column D = 4
        
        # Step 6: Add "record" and "<>" pattern to columns I:N starting at row 6
        pattern_values = ["record", "<>", "record", "<>", "record", "<>"]  # I, J, K, L, M, N
        
        for row_num in range(6, current_row + 1):  # Include the "<>" row
            for col_index, value in enumerate(pattern_values):
                col_num = 9 + col_index  # Column I = 9, J = 10, K = 11, L = 12, M = 13, N = 14
                question_ws.cell(row=row_num, column=col_num, value=value)
        
        # Step 7: Add percentage formula to column E starting at row 6
        # The denominator is the absolute reference to the cell in column D next to the "<>" in column C
        denominator_row = current_row  # This is the row with "<>" in column C
        
        for row_num in range(6, current_row + 1):  # Include the "<>" row
            # Create percentage formula: relative numerator / absolute denominator
            percentage_formula = f"=D{row_num}/$D${denominator_row}"
            cell = question_ws.cell(row=row_num, column=5, value=percentage_formula)  # Column E = 5
            # Format as percentage
            cell.number_format = '0.0%'
        
        logging.info(f"Added COUNTIFS formula to column D from row 6 to {current_row}")
        logging.info(f"Added record/<> pattern to columns I:N from row 6 to {current_row}")
        logging.info(f"Added percentage formula to column E from row 6 to {current_row} with denominator $D${denominator_row}")
        logging.info(f"Successfully extracted {len(response_rows)} response options for question {question_number}")
        
    except Exception as e:
        logging.error(f"Error extracting response options for question {question_number}: {e}")
        raise


def find_section_number_from_data_map(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the section number from column N of the data map in the same row as the first instance
    of the question number in column G.
    
    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)
        
    Returns:
        str: The section number from column N, or None if not found
    """
    try:
        # Search through column G starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            cell_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7
            
            if cell_value is not None:
                cell_str = str(cell_value).strip()
                
                # Check if this matches our question number
                if cell_str == str(question_number):
                    # Found the question, get column N value from same row
                    section_number = data_map_ws.cell(row=row, column=14).value  # Column N = 14
                    if section_number:
                        return str(section_number).strip()
                    else:
                        return None
        
        return None
        
    except Exception as e:
        logging.error(f"Error finding section number for question {question_number}: {e}")
        return None


def find_other_specify_child_text(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_number: int) -> str:
    """
    Finds the "Other Specify Child" question text from column C of the data map.
    
    Searches for the first row where:
    - Column G matches the question number
    - Column H contains "0, Open Text, 0, 0, 0, 0, 0, 0, 0, Other Specify Child"
    
    Args:
        data_map_ws: The data map worksheet
        question_number: The question number to search for (1-10)
        
    Returns:
        str: The question text from column C, or None if not found
    """
    try:
        target_h_pattern = "0, Open Text, 0, 0, , 0, 0, 0, 0, Other Specify Child"
        
        # Search through all rows starting from row 4 (where data begins)
        for row in range(4, data_map_ws.max_row + 1):
            # Check column G for question number match
            column_g_value = data_map_ws.cell(row=row, column=7).value  # Column G = 7
            # Check column H for the specific pattern
            column_h_value = data_map_ws.cell(row=row, column=8).value  # Column H = 8
            
            if (column_g_value is not None and 
                str(column_g_value).strip() == str(question_number) and
                column_h_value is not None and 
                str(column_h_value).strip() == target_h_pattern):
                
                # Found the matching row, get column C value (question text)
                question_text = data_map_ws.cell(row=row, column=3).value  # Column C = 3
                if question_text:
                    logging.info(f"Found Other Specify Child text at row {row} for question {question_number}")
                    return str(question_text).strip()
                else:
                    logging.warning(f"Found matching row {row} but column C is empty")
                    return None
        
        logging.info(f"No Other Specify Child pattern found for question {question_number}")
        return None
        
    except Exception as e:
        logging.error(f"Error finding Other Specify Child text for question {question_number}: {e}")
        return None


def extract_bracketed_text(text: str) -> str:
    """
    Extracts the first bracketed text from a string.
    
    For example: "[S1r6oe]: In which region..." returns "S1r6oe"
    
    Args:
        text: The text containing bracketed content
        
    Returns:
        str: The text inside the first brackets, or None if no brackets found
    """
    try:
        import re
        # Find the first occurrence of text within square brackets
        match = re.search(r'\[([^\]]+)\]', text)
        if match:
            return match.group(1)  # Return the content inside brackets
        else:
            return None
            
    except Exception as e:
        logging.error(f"Error extracting bracketed text from '{text}': {e}")
        return None


def save_processed_excel(workbook: openpyxl.Workbook, output_path: str) -> None:
    """
    Saves the processed workbook to the specified output path.

    Args:
        workbook (openpyxl.Workbook): The workbook to save.
        output_path (str): Path where to save the processed workbook.
        
    Raises:
        PermissionError: If unable to write to the file.
        OSError: If there's an OS-related error during saving.
    """
    try:
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        workbook.save(output_path)
        logging.info(f"Processed workbook saved successfully to: {output_path}")
    except Exception as e:
        logging.error(f"Error saving processed workbook to {output_path}: {e}")
        raise


def process_excel_file(input_path: str, output_path: str) -> None:
    """
    Main processing function that loads a raw Excel file and outputs a processed version.

    Args:
        input_path (str): Path to the input Excel file.
        output_path (str): Path for the output processed file.
        
    Raises:
        FileNotFoundError: If the input file doesn't exist.
        ValueError: If the input file is not a valid Excel file.
    """
    try:
        logging.info(f"Starting processing of {input_path}")
        
        # Load the raw Excel file
        workbook = load_raw_excel_file(input_path)
        
        # Perform initial setup
        initial_set_up(workbook)
        
        # Perform raw data initial setup
        raw_data_initial_setup(workbook)
        
        # Perform data map initial setup
        data_map_initial_setup(workbook)
        
        # Perform column question map initial setup
        column_question_map_initial_setup(workbook)
        
        # Save intermediate workbook with formulas (for debugging)
        base_path = Path(output_path)
        # Extract version number from output filename to keep consistent naming
        stem = base_path.stem
        if 'v' in stem:
            version_part = stem[stem.rfind('v'):]  # Gets 'v58' part
            formula_output_path = base_path.parent / f"test_processed_pilot{version_part}_formulas{base_path.suffix}"
        else:
            formula_output_path = base_path.parent / f"{stem}_formulas{base_path.suffix}"
        save_processed_excel(workbook, str(formula_output_path))
        logging.info(f"Saved intermediate workbook with formulas: {formula_output_path}")
        
        # Save workbook before Excel calculation
        save_processed_excel(workbook, output_path)
        
        # Calculate all formulas using Excel
        calculate_excel_formulas(output_path)
        
        # Reload workbook after Excel calculation with data_only=True to get calculated values
        workbook = openpyxl.load_workbook(output_path, data_only=True)
        
        # Perform question cutting
        question_cutting_processor(workbook)
        
        # Final save of the processed file
        save_processed_excel(workbook, output_path)
        
        logging.info(f"Successfully processed {input_path} -> {output_path}")
        
    except Exception as e:
        logging.error(f"Error processing Excel file {input_path}: {e}")
        raise


def get_next_version_filename(base_name: str = "test_processed_pilot") -> str:
    """
    Gets the next versioned filename in the output directory.
    
    Args:
        base_name (str): Base name for the file (without extension)
        
    Returns:
        str: Full path to the next versioned file
    """
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    
    version = 1
    while True:
        filename = f"{base_name}v{version}.xlsx"
        filepath = output_dir / filename
        if not filepath.exists():
            return str(filepath)
        version += 1


def main():
    """
    Main entry point for the survey processor v4.
    """
    import argparse
    
    logging.basicConfig(
        level=logging.INFO, 
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    parser = argparse.ArgumentParser(
        description='Process survey Excel files with improved functionality'
    )
    parser.add_argument('input_file', help='Path to the input Excel file')
    parser.add_argument('--output_file', help='Path for the output processed file (optional, auto-versions if not provided)')
    
    args = parser.parse_args()
    
    # Auto-generate versioned output filename if not provided
    output_file = args.output_file if args.output_file else get_next_version_filename()
    
    try:
        process_excel_file(args.input_file, output_file)
        print(f"Successfully processed: {args.input_file} -> {output_file}")
    except Exception as e:
        print(f"Error: {e}")
        return 1
    
    return 0


if __name__ == '__main__':
    exit(main())