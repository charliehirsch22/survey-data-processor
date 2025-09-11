#!/usr/bin/env python3
"""
Survey Data Processor

Converts raw survey data files into formatted Excel workbooks.

"""

import pandas as pd
import numpy as np
from pathlib import Path
import chardet
import logging
from typing import Dict, List, Optional, Tuple
import json
import re
from enum import Enum
import openpyxl
from openpyxl.utils import get_column_letter

def _is_sequential(numbers):
    """
    Helper function to check if a list of numbers forms a sequential pattern (like 1, 2, 3, 4...).
    
    Args:
        numbers (list): List of integers to check
        
    Returns:
        bool: True if numbers are sequential, False otherwise
    """
    if len(numbers) < 2:
        return False
    
    sorted_numbers = sorted(numbers)
    for i in range(1, len(sorted_numbers)):
        if sorted_numbers[i] - sorted_numbers[i-1] != 1:
            return False
    return True

def load_raw_survey(file_path: str) -> openpyxl.Workbook:
    """
    Loads an Excel file and returns an openpyxl Workbook object for editing.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        openpyxl.Workbook: The loaded workbook.
    
    Raises:
        FileNotFoundError: If the file doesn't exist.
        ValueError: If the file is not a valid Excel file.
    """
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if file_path_obj.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
            raise ValueError(f"Not a valid Excel file: {file_path}")
        
        return openpyxl.load_workbook(file_path)
    except Exception as e:
        logging.error(f"Error loading workbook {file_path}: {e}")
        raise

def rename_datamap_tab(workbook: openpyxl.Workbook) -> None:
    """
    Renames any worksheet with a name similar to 'datamap' to 'data map'.

    Args:
        workbook (openpyxl.Workbook): The workbook to edit.
    """
    for ws in workbook.worksheets:
        if re.search(r'data\s*map|datamap', ws.title, re.IGNORECASE):
            ws.title = 'data map'

def rename_rawdata_tab(workbook: openpyxl.Workbook) -> None:
    """
    Renames any worksheet with a name similar to 'A1' or 'raw data' to 'raw data'.

    Args:
        workbook (openpyxl.Workbook): The workbook to edit.
    """
    for ws in workbook.worksheets:
        if re.search(r'^A1$', ws.title, re.IGNORECASE) or re.search(r'raw\s*data', ws.title, re.IGNORECASE):
            ws.title = 'raw data'

def shift_worksheet_to_C2(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Ensures worksheet has blank first row and first two columns (A, B) with width 3,
    without removing any existing data. Data is shifted if necessary.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to modify.
    """
    # Check if first row has any data
    first_row_has_data = any(cell.value is not None for cell in ws[1])
    
    # Check if columns A or B have any data
    col_a_has_data = any(cell.value is not None for cell in ws['A'])
    col_b_has_data = any(cell.value is not None for cell in ws['B'])
    
    # Insert blank row at top if first row has data
    if first_row_has_data:
        ws.insert_rows(1)
    
    # Insert blank columns if A or B have data
    cols_to_insert = 0
    if col_b_has_data:
        cols_to_insert = 2
    elif col_a_has_data:
        cols_to_insert = 1
    
    if cols_to_insert > 0:
        ws.insert_cols(1, amount=cols_to_insert)
    
    # Always set width of columns A and B to 3
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 3
    
    # Remove gridlines
    ws.sheet_view.showGridLines = False

def process_raw_data_tab(workbook: openpyxl.Workbook) -> None:
    """
    Processes the 'raw data' tab by applying shift_worksheet_to_C2.

    Args:
        workbook (openpyxl.Workbook): The workbook containing the 'raw data' tab.
    """
    try:
        raw_data_ws = workbook['raw data']
        shift_worksheet_to_C2(raw_data_ws)
    except KeyError:
        logging.warning("No 'raw data' tab found in workbook")

def process_data_map_tab(workbook: openpyxl.Workbook) -> None:
    """
    Processes the 'data map' tab by applying shift_worksheet_to_C2 and adding column headers.

    Args:
        workbook (openpyxl.Workbook): The workbook containing the 'data map' tab.
    """
    try:
        data_map_ws = workbook['data map']
        
        # Apply the standard formatting first (this handles data preservation)
        shift_worksheet_to_C2(data_map_ws)
        
        # Now insert an additional row specifically for our headers
        # This ensures we don't overwrite any data that might be in row 2
        data_map_ws.insert_rows(2)
        
        # Add column headers in row 2 (which is now guaranteed to be empty)
        data_map_ws['C2'] = "Question Info"
        data_map_ws['D2'] = "Number Response" 
        data_map_ws['E2'] = "Text Map"
        data_map_ws['F2'] = "Helper #1"
        data_map_ws['G2'] = "Question Tag #1"
        data_map_ws['H2'] = "Question Tag #2"
        data_map_ws['I2'] = "Question Tag #3"
        data_map_ws['J2'] = "Question Tag #4"
        data_map_ws['K2'] = "Other Flag"
        
        # Format column C: width 45 and text wrap
        data_map_ws.column_dimensions['C'].width = 45
        for row in data_map_ws.iter_rows(min_col=3, max_col=3):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        # Autofit columns D and E
        for col in ['D', 'E']:
            max_length = 0
            column_cells = data_map_ws[col]
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
            data_map_ws.column_dimensions[col].width = adjusted_width
        
        # Center align column D
        for row in data_map_ws.iter_rows(min_col=4, max_col=4):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Fill Helper #1 column (F) with incrementing section numbers starting at F3 and after blank rows
        max_row = data_map_ws.max_row
        section_counter = 1
        
        # Always put 'Section #1' in F3
        data_map_ws['F3'] = f'Section #{section_counter}'
        section_counter += 1
        
        # Check each row starting from row 4 to find blank rows and fill the row after
        for row_num in range(4, max_row + 1):
            # Check if the current row is blank (all cells in the row are empty)
            row_is_blank = True
            for col_num in range(1, data_map_ws.max_column + 1):
                cell_value = data_map_ws.cell(row=row_num, column=col_num).value
                if cell_value is not None and str(cell_value).strip():
                    row_is_blank = False
                    break
            
            # If current row is blank and there's a next row, put next section number in F of next row
            if row_is_blank and row_num + 1 <= max_row:
                data_map_ws.cell(row=row_num + 1, column=6).value = f'Section #{section_counter}'  # Column F is column 6
                section_counter += 1
        
        # Fill Question Tag #1 column (G) based on rules
        system_keywords = ["[record]", "[uuid]", "[date]", "[markers]", "[status]"]
        
        for row_num in range(3, max_row + 1):
            # Check if this row has a section number in column F
            helper_cell = data_map_ws.cell(row=row_num, column=6).value  # Column F
            if helper_cell and str(helper_cell).startswith('Section #'):
                # Get the text in column C (Question Info)
                question_info = data_map_ws.cell(row=row_num, column=3).value  # Column C
                
                if question_info:
                    question_text = str(question_info).lower()
                    # Check if any system keywords are in the question text
                    is_system = any(keyword.lower() in question_text for keyword in system_keywords)
                    
                    if is_system:
                        data_map_ws.cell(row=row_num, column=7).value = "SYSTEM"  # Column G
                    else:
                        # Check the row below for patterns
                        next_row_question_info = data_map_ws.cell(row=row_num + 1, column=3).value  # Column C, next row
                        is_numerical = False
                        is_open_text = False
                        
                        if next_row_question_info:
                            next_row_text = str(next_row_question_info).lower()
                            
                            # Check for open text response
                            if "open text response" in next_row_text:
                                is_open_text = True
                            
                            # Check for numerical pattern - has "Values" AND no enumeration in Column D
                            if "values" in next_row_text:
                                # Find which section this row belongs to
                                current_section_start = row_num
                                next_section_start = max_row + 1
                                
                                # Find the next section boundary
                                for search_row in range(row_num + 1, max_row + 1):
                                    search_helper_cell = data_map_ws.cell(row=search_row, column=6).value
                                    if search_helper_cell and str(search_helper_cell).startswith('Section #'):
                                        next_section_start = search_row
                                        break
                                
                                # Check if there's any enumeration in Column D within the section
                                has_enumeration = False
                                for search_row in range(current_section_start + 1, next_section_start):
                                    col_d_value = data_map_ws.cell(row=search_row, column=4).value  # Column D
                                    if col_d_value:
                                        col_d_text = str(col_d_value).strip()
                                        # If we find any simple numbers (like "1", "2", "3"), it's enumerated
                                        if col_d_text.isdigit():
                                            has_enumeration = True
                                            break
                                
                                # If "Values" is present but NO enumeration in Column D, it's Numerical
                                if not has_enumeration:
                                    is_numerical = True
                        
                        if is_open_text:
                            data_map_ws.cell(row=row_num, column=7).value = "Open Text"  # Column G
                        elif is_numerical:
                            data_map_ws.cell(row=row_num, column=7).value = "Numerical"  # Column G
                        else:
                            data_map_ws.cell(row=row_num, column=7).value = "Select"  # Column G
                else:
                    data_map_ws.cell(row=row_num, column=7).value = "Select"  # Column G
        
        # Fill Question Tag #2 column (H) based on rules
        # Get raw data tab headers for comparison
        raw_data_headers = []
        try:
            raw_data_ws = workbook['raw data']
            for cell in raw_data_ws[2]:  # Row 2 contains headers
                if cell.value:
                    raw_data_headers.append(str(cell.value))
        except KeyError:
            logging.warning("No 'raw data' tab found for header comparison")
        
        for row_num in range(3, max_row + 1):
            # Check if Question Tag #1 (column G) is "Select"
            question_tag_1 = data_map_ws.cell(row=row_num, column=7).value  # Column G
            if question_tag_1 == "Select":
                # Get the Question Info text (column C)
                question_info = data_map_ws.cell(row=row_num, column=3).value  # Column C
                
                if question_info:
                    question_text_original = str(question_info).strip()
                    question_text = question_text_original.lower()
                    
                    # Check for Rank pattern first
                    if "rank" in question_text:
                        # Find section boundaries for this row
                        current_section_start = row_num
                        next_section_start = max_row + 1
                        
                        # Find the next section boundary
                        for search_row in range(row_num + 1, max_row + 1):
                            search_helper_cell = data_map_ws.cell(row=search_row, column=6).value
                            if search_helper_cell and str(search_helper_cell).startswith('Section #'):
                                next_section_start = search_row
                                break
                        
                        # Count enumerated values and bracketed text in Column D within section
                        enumerated_values = []
                        bracketed_text_rows = []
                        
                        for search_row in range(current_section_start + 1, next_section_start):
                            col_d_value = data_map_ws.cell(row=search_row, column=4).value  # Column D
                            if col_d_value:
                                col_d_text = str(col_d_value).strip()
                                # Check if it's a simple number (enumerated value)
                                if col_d_text.isdigit():
                                    enumerated_values.append(int(col_d_text))
                                # Check if it has brackets (funkier text)
                                elif '[' in col_d_text and ']' in col_d_text:
                                    bracketed_text_rows.append(col_d_text)
                        
                        # If enumerated values count equals bracketed text count, it's a Rank
                        if len(enumerated_values) > 0 and len(enumerated_values) == len(bracketed_text_rows):
                            data_map_ws.cell(row=row_num, column=8).value = "Rank"  # Column H
                        # Otherwise check for Single Select pattern
                        elif question_text.startswith('['):
                            # Extract text after the first open bracket until the closing bracket
                            bracket_end = question_text.find(']')
                            if bracket_end != -1:
                                bracket_content = question_text[1:bracket_end]
                                
                                # Count exact matches in raw data headers
                                exact_matches = sum(1 for header in raw_data_headers if header == bracket_content)
                                
                                if exact_matches == 1:
                                    data_map_ws.cell(row=row_num, column=8).value = "Single Select"  # Column H
                    # Check for other patterns if not rank  
                    else:
                        # Find section boundaries for this row
                        current_section_start = row_num
                        next_section_start = max_row + 1
                        
                        # Find the next section boundary
                        for search_row in range(row_num + 1, max_row + 1):
                            search_helper_cell = data_map_ws.cell(row=search_row, column=6).value
                            if search_helper_cell and str(search_helper_cell).startswith('Section #'):
                                next_section_start = search_row
                                break
                        
                        # Analyze Column D in section for enumerated values and brackets
                        enumerated_values = []
                        bracketed_text_rows = []
                        
                        for search_row in range(current_section_start + 1, next_section_start):
                            col_d_value = data_map_ws.cell(row=search_row, column=4).value  # Column D
                            if col_d_value:
                                col_d_text = str(col_d_value).strip()
                                # Check if it's a simple number (enumerated value)
                                if col_d_text.isdigit():
                                    enumerated_values.append(int(col_d_text))
                                # Check if it has brackets (funkier text)
                                elif '[' in col_d_text and ']' in col_d_text:
                                    bracketed_text_rows.append(col_d_text)
                        
                        # Extract bracket content for header matching
                        bracket_end = question_text_original.find(']')
                        bracket_content = ""
                        exact_matches = 0
                        if bracket_end != -1:
                            bracket_content = question_text_original[1:bracket_end]
                            exact_matches = sum(1 for header in raw_data_headers if header == bracket_content)
                        
                        # Determine question type based on patterns
                        if len(enumerated_values) > 0 and len(enumerated_values) == len(bracketed_text_rows):
                            # Equal enumerated and bracketed = Rank (but we already checked for "rank" keyword above)
                            pass  # This shouldn't happen since we checked "rank" first
                        elif len(enumerated_values) > 0 and len(bracketed_text_rows) > 0:
                            data_map_ws.cell(row=row_num, column=8).value = "Matrix"  # Column H
                        elif len(enumerated_values) > 0 and len(bracketed_text_rows) == 0 and exact_matches == 1 and question_text_original.startswith('['):
                            data_map_ws.cell(row=row_num, column=8).value = "Single Select"  # Column H
        
        # Fill Other Flag column (K) based on rules
        # Find all section boundaries first
        section_boundaries = {}  # {section_number: row_number}
        for row_num in range(3, max_row + 1):
            helper_cell = data_map_ws.cell(row=row_num, column=6).value  # Column F
            if helper_cell and str(helper_cell).startswith('Section #'):
                section_num = str(helper_cell).replace('Section #', '').strip()
                section_boundaries[section_num] = row_num
        
        for row_num in range(3, max_row + 1):
            # Check if Question Tag #1 (column G) is "Select"
            question_tag_1 = data_map_ws.cell(row=row_num, column=7).value  # Column G
            if question_tag_1 == "Select":
                # Find which section this row belongs to
                current_section = None
                current_section_start = None
                next_section_start = None
                
                # Find the current section
                for section_num, section_row in section_boundaries.items():
                    if section_row == row_num:
                        current_section = section_num
                        current_section_start = row_num
                        break
                
                if current_section and current_section_start:
                    # Find the next section to determine boundaries
                    section_numbers = sorted([int(s) for s in section_boundaries.keys()])
                    current_section_int = int(current_section)
                    
                    if current_section_int < max(section_numbers):
                        next_section_int = current_section_int + 1
                        next_section_start = section_boundaries[str(next_section_int)]
                    else:
                        next_section_start = max_row + 1
                    
                    # Search for "Other (please specify)" in column E within section boundaries
                    for search_row in range(current_section_start, next_section_start):
                        text_map_cell = data_map_ws.cell(row=search_row, column=5).value  # Column E
                        if text_map_cell and "other (please specify)" in str(text_map_cell).lower():
                            data_map_ws.cell(row=row_num, column=11).value = "Other Flag"  # Column K
                            break
        
        # Add question numbering in Column B
        question_counter = 1
        
        for row_num in range(3, max_row + 1):
            # Check if this row has a section number (indicates a potential new question)
            helper_cell = data_map_ws.cell(row=row_num, column=6).value  # Column F (Helper #1)
            question_tag_1 = data_map_ws.cell(row=row_num, column=7).value  # Column G (Question Tag #1)
            other_flag = data_map_ws.cell(row=row_num, column=11).value  # Column K (Other Flag)
            question_text = data_map_ws.cell(row=row_num, column=3).value  # Column C (Question Info)
            
            if helper_cell and str(helper_cell).startswith('Section #'):
                # Skip SYSTEM questions - they don't get question numbers
                if question_tag_1 == "SYSTEM":
                    continue
                
                # Check if this is an "Other" continuation section (not a new question)
                is_other_continuation = False
                
                # Method 1: Check if question text contains "oe]" indicating other entry
                if question_text and "oe]" in str(question_text).lower():
                    is_other_continuation = True
                
                # Method 2: Check if previous question had "Other Flag"
                if not is_other_continuation:
                    # Look back to find the previous section with a question
                    for prev_row in range(row_num - 1, 2, -1):  # Go backwards from current row
                        prev_helper_cell = data_map_ws.cell(row=prev_row, column=6).value
                        if prev_helper_cell and str(prev_helper_cell).startswith('Section #'):
                            prev_other_flag = data_map_ws.cell(row=prev_row, column=11).value
                            if prev_other_flag == "Other Flag":
                                is_other_continuation = True
                            break
                
                if not is_other_continuation:
                    # This is a genuine new question - assign question number
                    data_map_ws.cell(row=row_num, column=2).value = f"Q{question_counter}"  # Column B
                    question_counter += 1
                else:
                    # This is an "Other" continuation - give it the same number as the previous question
                    prev_question_num = f"Q{question_counter - 1}" if question_counter > 1 else ""
                    data_map_ws.cell(row=row_num, column=2).value = prev_question_num  # Column B
        
        # Process individual questions - start with Q1
        process_question_cutting(workbook, data_map_ws, max_row)
        
    except KeyError:
        logging.warning("No 'data map' tab found in workbook")

def process_question_cutting(workbook: openpyxl.Workbook, data_map_ws: openpyxl.worksheet.worksheet.Worksheet, max_row: int) -> None:
    """
    Process question cutting by finding questions and creating individual tabs.
    
    Args:
        workbook (openpyxl.Workbook): The workbook to add question tabs to.
        data_map_ws (openpyxl.worksheet.worksheet.Worksheet): The data map worksheet.
        max_row (int): Maximum row number in the data map.
    """
    # Find Q1 in the data map
    q1_info = find_question_info(data_map_ws, "Q1", max_row)
    
    if q1_info:
        # Create Q1 tab and process it
        create_question_tab(workbook, "Q1", q1_info)
    else:
        logging.warning("Q1 not found in data map")

def find_question_info(data_map_ws: openpyxl.worksheet.worksheet.Worksheet, question_num: str, max_row: int) -> dict:
    """
    Find information about a specific question in the data map.
    
    Args:
        data_map_ws: The data map worksheet
        question_num: Question number to find (e.g., "Q1")
        max_row: Maximum row to search
        
    Returns:
        dict: Question information including row, question type, tags, etc.
    """
    for row_num in range(3, max_row + 1):
        question_number = data_map_ws.cell(row=row_num, column=2).value  # Column B
        
        if question_number == question_num:
            question_info = {
                'row': row_num,
                'question_number': question_num,
                'question_text': data_map_ws.cell(row=row_num, column=3).value,  # Column C
                'tag_1': data_map_ws.cell(row=row_num, column=7).value,  # Column G
                'tag_2': data_map_ws.cell(row=row_num, column=8).value,  # Column H
                'tag_3': data_map_ws.cell(row=row_num, column=9).value,  # Column I
                'tag_4': data_map_ws.cell(row=row_num, column=10).value,  # Column J
                'other_flag': data_map_ws.cell(row=row_num, column=11).value,  # Column K
                'helper_1': data_map_ws.cell(row=row_num, column=6).value,  # Column F
            }
            
            logging.info(f"Found {question_num}: Tag1={question_info['tag_1']}, Tag2={question_info['tag_2']}, Other Flag={question_info['other_flag']}")
            return question_info
    
    return None

def create_question_tab(workbook: openpyxl.Workbook, tab_name: str, question_info: dict) -> None:
    """
    Create a new tab for a question and populate it based on question type.
    
    Args:
        workbook: The workbook to add the tab to
        tab_name: Name of the new tab (e.g., "Q1")
        question_info: Dictionary containing question information
    """
    # Create new worksheet
    question_ws = workbook.create_sheet(title=tab_name)
    
    # Determine which helper function to call based on question tags
    tag_1 = question_info.get('tag_1', '')
    tag_2 = question_info.get('tag_2', '')
    other_flag = question_info.get('other_flag', '')
    
    logging.info(f"Creating {tab_name} tab with Tag1='{tag_1}', Tag2='{tag_2}'")
    
    # Route to specific helper function for exact tag combination
    if tag_1 == "Select" and tag_2 == "Single Select" and other_flag == "Other Flag":
        cut_select_singleselect_otherflag(question_ws, question_info)
    elif tag_1 == "Select" and tag_2 == "Single Select":
        cut_single_select_question(question_ws, question_info)
    elif tag_1 == "Select" and tag_2 == "Matrix":
        cut_matrix_question(question_ws, question_info)
    elif tag_1 == "Select" and tag_2 == "Rank":
        cut_rank_question(question_ws, question_info)
    elif tag_1 == "Numerical":
        cut_numerical_question(question_ws, question_info)
    elif tag_1 == "Open Text":
        cut_open_text_question(question_ws, question_info)
    else:
        cut_default_question(question_ws, question_info)

# Specific helper functions for exact question type combinations
def cut_select_singleselect_otherflag(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_info: dict) -> None:
    """Helper function for Select + Single Select + Other Flag questions."""
    logging.info(f"Cutting Select + Single Select + Other Flag question: {question_info['question_number']}")
    
    # Apply the formatting function to make first 2 columns blank with width 3 and blank row
    shift_worksheet_to_C2(question_ws)
    
    # Put the question text from Column C of the data map into cell C2 of the new worksheet
    question_text = question_info.get('question_text', '')
    if question_text:
        question_ws['C2'] = question_text
        logging.info(f"Added question text to C2: '{str(question_text)[:50]}'")

# Placeholder helper functions for different question types
def cut_single_select_question(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_info: dict) -> None:
    """Helper function for Single Select questions."""
    logging.info(f"Cutting Single Select question: {question_info['question_number']}")
    # Placeholder - will be implemented based on your specifications

def cut_matrix_question(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_info: dict) -> None:
    """Helper function for Matrix questions."""
    logging.info(f"Cutting Matrix question: {question_info['question_number']}")
    # Placeholder - will be implemented based on your specifications

def cut_rank_question(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_info: dict) -> None:
    """Helper function for Rank questions."""
    logging.info(f"Cutting Rank question: {question_info['question_number']}")
    # Placeholder - will be implemented based on your specifications

def cut_numerical_question(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_info: dict) -> None:
    """Helper function for Numerical questions."""
    logging.info(f"Cutting Numerical question: {question_info['question_number']}")
    # Placeholder - will be implemented based on your specifications

def cut_open_text_question(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_info: dict) -> None:
    """Helper function for Open Text questions."""
    logging.info(f"Cutting Open Text question: {question_info['question_number']}")
    # Placeholder - will be implemented based on your specifications

def cut_default_question(question_ws: openpyxl.worksheet.worksheet.Worksheet, question_info: dict) -> None:
    """Default helper function for unrecognized question types."""
    logging.info(f"Cutting default question: {question_info['question_number']} (Tag1={question_info['tag_1']}, Tag2={question_info['tag_2']})")
    # Placeholder - will be implemented based on your specifications

def save_workbook(workbook: openpyxl.Workbook, file_path: str) -> None:
    """
    Saves the workbook to the specified file path.

    Args:
        workbook (openpyxl.Workbook): The workbook to save.
        file_path (str): Path where to save the workbook.
        
    Raises:
        PermissionError: If unable to write to the file.
        OSError: If there's an OS-related error during saving.
    """
    try:
        workbook.save(file_path)
        logging.info(f"Workbook saved successfully to: {file_path}")
    except Exception as e:
        logging.error(f"Error saving workbook to {file_path}: {e}")
        raise

def process_survey_file(input_path: str, output_path: Optional[str] = None) -> None:
    """
    Processes a survey Excel file by renaming tabs, formatting them, and saving the result.

    Args:
        input_path (str): Path to the input Excel file.
        output_path (Optional[str]): Path for the output file. If None, overwrites input.
        
    Raises:
        FileNotFoundError: If the input file doesn't exist.
        ValueError: If the input file is not a valid Excel file.
    """
    try:
        workbook = load_raw_survey(input_path)
        rename_datamap_tab(workbook)
        rename_rawdata_tab(workbook)
        process_raw_data_tab(workbook)
        process_data_map_tab(workbook)
        
        save_path = output_path if output_path else input_path
        save_workbook(workbook, save_path)
        
    except Exception as e:
        logging.error(f"Error processing survey file {input_path}: {e}")
        raise

def main():
    """
    Main entry point for the survey processor.
    """
    import argparse
    
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    parser = argparse.ArgumentParser(description='Process survey Excel files by renaming tabs')
    parser.add_argument('input_file', help='Path to the input Excel file')
    parser.add_argument('-o', '--output', help='Path for the output file (optional, overwrites input if not specified)')
    
    args = parser.parse_args()
    
    try:
        process_survey_file(args.input_file, args.output)
        print(f"Successfully processed survey file: {args.input_file}")
        if args.output:
            print(f"Output saved to: {args.output}")
        else:
            print("Input file updated in place")
    except Exception as e:
        print(f"Error: {e}")
        return 1
    
    return 0

if __name__ == '__main__':
    exit(main())